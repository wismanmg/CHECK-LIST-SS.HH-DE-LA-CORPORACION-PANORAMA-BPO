# -*- coding: utf-8 -*-
"""
Migra los casos del Excel local a Supabase (una sola vez).

Uso:
  1. Ejecuta antes supabase_tablas.sql en el SQL Editor de Supabase.
  2. Completa SUPABASE_URL y SUPABASE_KEY aquí abajo (o define las variables de entorno).
  3. Doble clic en MIGRAR A SUPABASE.bat  (o: python migrar_a_supabase.py)

El script no borra nada del Excel; solo copia. Si se ejecuta dos veces,
duplicaría los casos: hazlo una sola vez sobre una base vacía.
"""
import json
import os
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

# Credenciales del proyecto (Supabase -> Settings -> API Keys)
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://aalxjzbemeacdpqqtfth.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "sb_secret_9My0gYbM_myYd3uhLhavRw_gO_Fmr7b")

CARPETA = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_EXCEL = os.path.join(CARPETA, "CHECK LIST SS.HH DE LA CORPORACION (Respuestas).xlsx")
ZONA = timezone(timedelta(hours=-5))  # Perú


def _sb(metodo, tabla, datos):
    peticion = urllib.request.Request(
        f"{SUPABASE_URL.rstrip('/')}/rest/v1/{tabla}",
        data=json.dumps(datos).encode("utf-8"), method=metodo,
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
                 "Content-Type": "application/json", "Prefer": "return=minimal"})
    try:
        with urllib.request.urlopen(peticion, timeout=30) as r:
            r.read()
    except urllib.error.HTTPError as e:
        raise SystemExit(f"ERROR de Supabase ({e.code}): {e.read().decode('utf-8', 'ignore')[:400]}")


def texto(v):
    return "" if v is None else str(v).strip()


def main():
    if "TU-PROYECTO" in SUPABASE_URL or "PEGA-AQUI" in SUPABASE_KEY:
        print("Primero completa SUPABASE_URL y SUPABASE_KEY dentro de este archivo.")
        input("Presiona Enter para salir...")
        return

    # Protección: si la nube ya tiene casos, no volver a migrar (evita duplicados)
    peticion = urllib.request.Request(
        f"{SUPABASE_URL.rstrip('/')}/rest/v1/observaciones?select=id&limit=1",
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"})
    with urllib.request.urlopen(peticion, timeout=30) as r:
        if json.loads(r.read().decode()):
            print("La base de Supabase YA tiene casos registrados.")
            print("No se migró nada para evitar duplicados. (Esta migración se hace una sola vez.)")
            input("Presiona Enter para salir...")
            return

    ws = load_workbook(ARCHIVO_EXCEL)["Respuestas de formulario 1"]
    casos = []
    for fila in ws.iter_rows(min_row=2, max_col=20, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in fila):
            continue
        fecha = fila[0]
        marca = (fecha.replace(tzinfo=ZONA).isoformat()
                 if isinstance(fecha, datetime) else datetime.now(ZONA).isoformat())
        monto = None
        if fila[18] is not None and str(fila[18]).strip() != "":
            try:
                monto = float(str(fila[18]).replace(",", "."))
            except ValueError:
                monto = None
        casos.append({
            "marca_temporal": marca,
            "edificio": texto(fila[1]), "piso": texto(fila[2]), "ubicacion": texto(fila[3]),
            "empresas": texto(fila[4]), "sshh": texto(fila[5]), "lavatorio": texto(fila[6]),
            "mesa": texto(fila[7]), "inodoro": texto(fila[8]), "puertas": texto(fila[9]),
            "urinario": texto(fila[10]), "descripcion": texto(fila[11]),
            "comentario": texto(fila[12]), "dispensador": texto(fila[13]),
            "comentario_fmi": texto(fila[14]), "proveedor": texto(fila[15]),
            "estatus": texto(fila[16]), "po": texto(fila[17]),
            "monto": monto, "prioridad": texto(fila[19]).upper(),
        })

    print(f"Casos encontrados en el Excel: {len(casos)}")
    # marca_temporal tiene default now(): para conservar la fecha original hay que
    # enviarla explícitamente, cosa que ya hacemos arriba.
    LOTE = 50
    for i in range(0, len(casos), LOTE):
        _sb("POST", "observaciones", casos[i:i + LOTE])
        print(f"  subidos {min(i + LOTE, len(casos))}/{len(casos)}")
    print("LISTO: Migracion completada. Revisa la tabla 'observaciones' en Supabase.")
    input("Presiona Enter para salir...")


if __name__ == "__main__":
    main()
