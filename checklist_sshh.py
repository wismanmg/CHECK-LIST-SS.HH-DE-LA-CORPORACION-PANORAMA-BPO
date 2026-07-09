# -*- coding: utf-8 -*-
"""
CHECK LIST SS.HH DE LA CORPORACION - Aplicación web local.

Dos vistas:
  /         Formulario de registro (igual al Google Form original)
  /control  Panel de control: seguimiento de COTIZACION PROVEEDOR, ESTATUS y
            COTIZACION PO. Con MODO ADMINISTRADOR (clave) para editar cualquier
            campo de una observación registrada.

Todas las respuestas se guardan en el Excel existente
"CHECK LIST SS.HH DE LA CORPORACION (Respuestas).xlsx" (hoja "Respuestas de formulario 1"),
respetando las 18 columnas actuales.
"""
import base64
import io
import json
import os
import threading
import urllib.request
import webbrowser
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# CARPETA_DATOS permite mover la base de datos (usado por Docker: /datos)
CARPETA = os.environ.get("CARPETA_DATOS") or os.path.dirname(os.path.abspath(__file__))
ARCHIVO_EXCEL = os.path.join(CARPETA, "CHECK LIST SS.HH DE LA CORPORACION (Respuestas).xlsx")
HOJA = "Respuestas de formulario 1"
HOJA_HISTORIAL = "HISTORIAL"
# PORT la asignan plataformas en la nube (Render, Railway, Heroku); PUERTO es la local
PUERTO = int(os.environ.get("PORT") or os.environ.get("PUERTO") or "8740")
ABRIR_NAVEGADOR = os.environ.get("SIN_NAVEGADOR") != "1" and "PORT" not in os.environ
NUM_COLUMNAS = 20  # Marca temporal ... COTIZACION PO, MONTO (S/), PRIORIDAD

# >>> CLAVE DEL MODO ADMINISTRADOR: edita todo (cámbiala aquí o por variable de entorno) <<<
CLAVE_ADMIN = os.environ.get("CLAVE_ADMIN", "FMI2026")
# >>> CLAVE LIMITADA: solo permite editar COMENTARIO FMI / PROVEEDOR y COTIZACION PROVEEDOR <<<
CLAVE_EDITOR = os.environ.get("CLAVE_EDITOR", "FME2026")

ENCABEZADOS = [
    "Marca temporal", "EDIFICIO", "PISO", "UBICACION", "EMPRESAS", "SS.HH ",
    "LAVATORIO", "MESA DE LABATORIOS", "INODORO", "PUERTAS DE INODOROS",
    "URINARIO", "DESCRIPCION ", "Comentario", "DiSPENSADOR",
    "COMENTARIO FMI / PROVEEDOR", "COTIZACION  PROVEEDOR", "ESTATUS",
    "COTIZACION  PO", "MONTO (S/)", "PRIORIDAD",
]

# ------- Supabase (PostgreSQL en la nube). Si estas variables existen, la app
# ------- guarda ahí en lugar del Excel local. Ideal para Render (disco efímero).
# "".join(x.split()) elimina espacios y saltos de línea que se cuelan al pegar la clave
SUPABASE_URL = "".join(os.environ.get("SUPABASE_URL", "").split()).rstrip("/")
SUPABASE_KEY = "".join(os.environ.get("SUPABASE_KEY", "").split())
USAR_SUPABASE = bool(SUPABASE_URL and SUPABASE_KEY)
ZONA = timezone(timedelta(hours=int(os.environ.get("ZONA_HORARIA", "-5"))))  # Perú: UTC-5

ESTADOS = ["PENDIENTE", "SOLI.COTI.PROV", "COTIZACION", "APROBADO", "EN_ EJECUCION", "ATENDIDO"]

candado = threading.Lock()

# ---------------------------------------------------------------- Excel

def filas_con_datos(ws):
    """Devuelve [(nro_fila_excel, [18 valores])] solo de filas con contenido."""
    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=NUM_COLUMNAS, values_only=True), start=2):
        if any(v is not None and str(v).strip() != "" for v in row):
            filas.append((idx, list(row)))
    return filas


def siguiente_fila_libre(ws):
    datos = filas_con_datos(ws)
    return (datos[-1][0] + 1) if datos else 2


def _escribir(ws, fila, col, valor):
    celda = ws.cell(row=fila, column=col, value=valor)
    celda.font = Font(name="Arial", size=10)


def _hoja_historial(wb):
    """Devuelve la hoja HISTORIAL, creándola con formato si no existe."""
    if HOJA_HISTORIAL in wb.sheetnames:
        return wb[HOJA_HISTORIAL]
    h = wb.create_sheet(HOJA_HISTORIAL)
    h.append(["FECHA Y HORA", "FILA DEL CASO", "USUARIO", "CAMPO", "VALOR ANTERIOR", "VALOR NUEVO"])
    for c in h[1]:
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="673AB7")
    for letra, ancho in zip("ABCDEF", [19, 13, 18, 26, 40, 40]):
        h.column_dimensions[letra].width = ancho
    h.freeze_panes = "A2"
    return h


def _log(hist, fila_caso, usuario, campo, antes, despues):
    hist.append([datetime.now().strftime("%d/%m/%Y %H:%M:%S"), fila_caso, usuario, campo, antes, despues])
    for c in hist[hist.max_row]:
        c.font = Font(name="Arial", size=9)


def guardar_registro_excel(d):
    with candado:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb[HOJA]
        fila = siguiente_fila_libre(ws)
        valores = {
            1: datetime.now(),
            2: d.get("edificio", ""),
            3: d.get("piso", ""),
            4: d.get("ubicacion", ""),
            5: d.get("empresas", ""),
            6: d.get("sshh", ""),
            7: d.get("lavatorio", ""),
            8: d.get("mesa_lavatorios", ""),
            9: d.get("inodoro", ""),
            10: d.get("puertas_inodoros", ""),
            11: d.get("urinario", ""),
            12: d.get("descripcion", ""),
            13: d.get("comentario", ""),
            17: "PENDIENTE",  # toda observación nueva nace con estatus PENDIENTE
            20: d.get("prioridad", ""),
        }
        for col, val in valores.items():
            _escribir(ws, fila, col, val)
        ws.cell(row=fila, column=1).number_format = "dd/mm/yyyy hh:mm:ss"
        _log(_hoja_historial(wb), fila, "FORMULARIO", "NUEVO REGISTRO", "",
             d.get("descripcion", "") or "(sin descripción)")
        wb.save(ARCHIVO_EXCEL)
        return fila


def _validar_fila(ws, d):
    fila = int(d["fila"])
    if fila < 2 or fila > ws.max_row:
        raise ValueError("Fila fuera de rango")
    return fila


def _aplicar_cambios(d, campos, usuario):
    """Escribe los campos recibidos y registra en HISTORIAL solo lo que realmente cambió."""
    with candado:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb[HOJA]
        fila = _validar_fila(ws, d)
        hist = None
        for col, clave in campos.items():
            if clave not in d:
                continue
            valor = str(d[clave]).strip()
            if clave == "monto" and valor:
                try:
                    valor = float(valor.replace(",", "."))
                except ValueError:
                    pass
            viejo = ws.cell(row=fila, column=col).value
            viejo = "" if viejo is None else viejo
            if str(viejo).strip() != str(valor).strip():
                if hist is None:
                    hist = _hoja_historial(wb)
                nombre = str(ws.cell(row=1, column=col).value or f"Columna {col}").strip()
                _log(hist, fila, usuario, nombre, str(viejo).strip(), str(valor).strip())
            _escribir(ws, fila, col, valor)
            if clave == "monto" and isinstance(valor, float):
                ws.cell(row=fila, column=col).number_format = "#,##0.00"
        wb.save(ARCHIVO_EXCEL)


def leer_historial_excel():
    with candado:
        wb = load_workbook(ARCHIVO_EXCEL)
        if HOJA_HISTORIAL not in wb.sheetnames:
            return []
        filas = []
        for row in wb[HOJA_HISTORIAL].iter_rows(min_row=2, max_col=6, values_only=True):
            if row[0] is None:
                continue
            filas.append([("" if v is None else str(v)) for v in row])
        return filas

# ---------------------------------------------------------------- Supabase

CAMPO_A_COLUMNA = {
    "edificio": "edificio", "piso": "piso", "ubicacion": "ubicacion", "empresas": "empresas",
    "sshh": "sshh", "lavatorio": "lavatorio", "mesa": "mesa", "inodoro": "inodoro",
    "puertas": "puertas", "urinario": "urinario", "descripcion": "descripcion",
    "comentario": "comentario", "dispensador": "dispensador", "prioridad": "prioridad",
    "comentario_mili": "comentario_fmi", "proveedor": "proveedor", "estatus": "estatus",
    "po": "po", "monto": "monto",
}
NOMBRE_CAMPO = {
    "edificio": "EDIFICIO", "piso": "PISO", "ubicacion": "UBICACION", "empresas": "EMPRESAS",
    "sshh": "SS.HH", "lavatorio": "LAVATORIO", "mesa": "MESA DE LABATORIOS",
    "inodoro": "INODORO", "puertas": "PUERTAS DE INODOROS", "urinario": "URINARIO",
    "descripcion": "DESCRIPCION", "comentario": "Comentario", "dispensador": "DiSPENSADOR",
    "comentario_mili": "COMENTARIO FMI / PROVEEDOR", "proveedor": "COTIZACION PROVEEDOR",
    "estatus": "ESTATUS", "po": "COTIZACION PO", "monto": "MONTO (S/)", "prioridad": "PRIORIDAD",
}


def _sb(metodo, tabla, datos=None, params=""):
    """Llamada REST a Supabase (PostgREST) usando solo la librería estándar."""
    url = f"{SUPABASE_URL}/rest/v1/{tabla}" + (("?" + params) if params else "")
    cuerpo = json.dumps(datos).encode("utf-8") if datos is not None else None
    peticion = urllib.request.Request(url, data=cuerpo, method=metodo, headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    })
    try:
        with urllib.request.urlopen(peticion, timeout=20) as r:
            texto = r.read().decode("utf-8")
            return json.loads(texto) if texto else []
    except urllib.error.HTTPError as e:
        detalle = e.read().decode("utf-8", "ignore")[:300]
        raise RuntimeError(f"Supabase respondió {e.code}: {detalle}")


def _fecha_local(iso):
    try:
        return datetime.fromisoformat(iso).astimezone(ZONA)
    except (ValueError, TypeError):
        return None


def guardar_registro_sb(d):
    fila = {
        "edificio": d.get("edificio", ""), "piso": d.get("piso", ""),
        "ubicacion": d.get("ubicacion", ""), "empresas": d.get("empresas", ""),
        "sshh": d.get("sshh", ""), "lavatorio": d.get("lavatorio", ""),
        "mesa": d.get("mesa_lavatorios", ""), "inodoro": d.get("inodoro", ""),
        "puertas": d.get("puertas_inodoros", ""), "urinario": d.get("urinario", ""),
        "descripcion": d.get("descripcion", ""), "comentario": d.get("comentario", ""),
        "prioridad": d.get("prioridad", ""), "estatus": "PENDIENTE",
    }
    creado = _sb("POST", "observaciones", fila)
    nuevo_id = creado[0]["id"]
    _sb("POST", "historial", {
        "caso_id": nuevo_id, "usuario": "FORMULARIO", "campo": "NUEVO REGISTRO",
        "valor_anterior": "", "valor_nuevo": d.get("descripcion", "") or "(sin descripción)",
    })
    return nuevo_id


def leer_registros_sb():
    filas = _sb("GET", "observaciones", params="select=*&order=id.asc")
    ahora = datetime.now(ZONA)
    registros = []
    for f in filas:
        fecha_txt, dias = "", None
        dt = _fecha_local(f.get("marca_temporal"))
        if dt:
            fecha_txt = dt.strftime("%d/%m/%Y %H:%M")
            dias = (ahora - dt).days
        registros.append({
            "fila": f["id"], "fecha": fecha_txt, "dias": dias,
            "edificio": f.get("edificio") or "", "piso": f.get("piso") or "",
            "ubicacion": f.get("ubicacion") or "", "empresas": f.get("empresas") or "",
            "sshh": f.get("sshh") or "", "lavatorio": f.get("lavatorio") or "",
            "mesa": f.get("mesa") or "", "inodoro": f.get("inodoro") or "",
            "puertas": f.get("puertas") or "", "urinario": f.get("urinario") or "",
            "descripcion": f.get("descripcion") or "", "comentario": f.get("comentario") or "",
            "dispensador": f.get("dispensador") or "",
            "comentario_mili": f.get("comentario_fmi") or "",
            "proveedor": f.get("proveedor") or "",
            "estatus": (f.get("estatus") or "").strip(),
            "po": f.get("po") or "",
            "monto": "" if f.get("monto") is None else str(f["monto"]),
            "prioridad": (f.get("prioridad") or "").strip().upper(),
        })
    return registros


def _aplicar_cambios_sb(d, claves, usuario):
    id_caso = int(d["fila"])
    actual = _sb("GET", "observaciones", params=f"select=*&id=eq.{id_caso}")
    if not actual:
        raise ValueError("Caso no encontrado")
    actual = actual[0]
    cambios, eventos = {}, []
    for clave in claves:
        if clave not in d:
            continue
        columna = CAMPO_A_COLUMNA[clave]
        nuevo = str(d[clave]).strip()
        if clave == "monto":
            valor = None
            if nuevo:
                try:
                    valor = float(nuevo.replace(",", "."))
                except ValueError:
                    valor = None
            viejo_s = "" if actual.get(columna) is None else str(actual[columna])
            nuevo_s = "" if valor is None else str(valor)
            nuevo_db = valor
        else:
            viejo_s = str(actual.get(columna) or "").strip()
            nuevo_s = nuevo
            nuevo_db = nuevo
        if viejo_s != nuevo_s:
            cambios[columna] = nuevo_db
            eventos.append({
                "caso_id": id_caso, "usuario": usuario,
                "campo": NOMBRE_CAMPO.get(clave, clave),
                "valor_anterior": viejo_s, "valor_nuevo": nuevo_s,
            })
    if cambios:
        _sb("PATCH", "observaciones", cambios, params=f"id=eq.{id_caso}")
        for evento in eventos:
            _sb("POST", "historial", evento)


def leer_historial_sb():
    filas = _sb("GET", "historial", params="select=*&order=id.asc")
    salida = []
    for f in filas:
        dt = _fecha_local(f.get("fecha"))
        salida.append([
            dt.strftime("%d/%m/%Y %H:%M:%S") if dt else "",
            str(f.get("caso_id") or ""), f.get("usuario") or "",
            f.get("campo") or "", f.get("valor_anterior") or "", f.get("valor_nuevo") or "",
        ])
    return salida

# ------------------------------------------------- despachadores de backend

def guardar_registro(d):
    return guardar_registro_sb(d) if USAR_SUPABASE else guardar_registro_excel(d)


def actualizar_control(d, usuario=""):
    """Actualiza solo las columnas de seguimiento."""
    if USAR_SUPABASE:
        _aplicar_cambios_sb(d, ["comentario_mili", "proveedor", "estatus", "po", "monto"], usuario)
    else:
        _aplicar_cambios(d, {15: "comentario_mili", 16: "proveedor", 17: "estatus", 18: "po", 19: "monto"}, usuario)


def editar_registro(d, usuario=""):
    """Modo administrador: edita todos los campos de la observación."""
    if USAR_SUPABASE:
        _aplicar_cambios_sb(d, ["edificio", "piso", "ubicacion", "empresas", "sshh", "lavatorio",
                                "mesa", "inodoro", "puertas", "urinario", "descripcion",
                                "comentario", "dispensador", "prioridad"], usuario)
    else:
        _aplicar_cambios(d, {
            2: "edificio", 3: "piso", 4: "ubicacion", 5: "empresas",
            6: "sshh", 7: "lavatorio", 8: "mesa", 9: "inodoro",
            10: "puertas", 11: "urinario", 12: "descripcion",
            13: "comentario", 14: "dispensador", 20: "prioridad",
        }, usuario)


def leer_historial():
    return leer_historial_sb() if USAR_SUPABASE else leer_historial_excel()


def leer_registros():
    if USAR_SUPABASE:
        return leer_registros_sb()
    return leer_registros_excel()


def leer_registros_excel():
    with candado:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb[HOJA]
        registros = []
        ahora = datetime.now()
        for fila, v in filas_con_datos(ws):
            v = [("" if x is None else x) for x in v]
            fecha = v[0]
            dias = None
            if isinstance(fecha, datetime):
                dias = (ahora - fecha).days
                fecha = fecha.strftime("%d/%m/%Y %H:%M")
            registros.append({
                "fila": fila,
                "fecha": str(fecha),
                "dias": dias,
                "edificio": str(v[1]),
                "piso": str(v[2]),
                "ubicacion": str(v[3]),
                "empresas": str(v[4]),
                "sshh": str(v[5]),
                "lavatorio": str(v[6]),
                "mesa": str(v[7]),
                "inodoro": str(v[8]),
                "puertas": str(v[9]),
                "urinario": str(v[10]),
                "descripcion": str(v[11]),
                "comentario": str(v[12]),
                "dispensador": str(v[13]),
                "comentario_mili": str(v[14]),
                "proveedor": str(v[15]),
                "estatus": str(v[16]).strip(),
                "po": str(v[17]),
                "monto": str(v[18]),
                "prioridad": str(v[19]).strip().upper(),
            })
        return registros

def generar_resumen():
    """Genera un Excel en memoria con hoja RESUMEN (totales) y hoja DETALLE (todos los casos).
    Funciona con ambos backends porque parte de leer_registros()."""
    registros = leer_registros()
    encabezados = ENCABEZADOS
    datos = []
    for r in registros:
        monto = ""
        if r["monto"]:
            try:
                monto = float(str(r["monto"]).replace(",", "."))
            except ValueError:
                monto = r["monto"]
        datos.append([
            r["fecha"], r["edificio"], r["piso"], r["ubicacion"], r["empresas"], r["sshh"],
            r["lavatorio"], r["mesa"], r["inodoro"], r["puertas"], r["urinario"],
            r["descripcion"], r["comentario"], r["dispensador"], r["comentario_mili"],
            r["proveedor"], r["estatus"], r["po"], monto, r["prioridad"],
        ])

    morado = PatternFill("solid", start_color="673AB7")
    gris = PatternFill("solid", start_color="E8EAED")
    f_titulo = Font(name="Arial", size=14, bold=True, color="673AB7")
    f_cab = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    f_sub = Font(name="Arial", size=11, bold=True)
    f_normal = Font(name="Arial", size=10)
    f_negrita = Font(name="Arial", size=10, bold=True)

    out = Workbook()

    # ---------------- hoja DETALLE ----------------
    det = out.active
    det.title = "DETALLE"
    for c, h in enumerate(encabezados, 1):
        celda = det.cell(row=1, column=c, value=h)
        celda.font = f_cab
        celda.fill = morado
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    for r, vals in enumerate(datos, start=2):
        for c, v in enumerate(vals, 1):
            celda = det.cell(row=r, column=c, value=v)
            celda.font = f_normal
            celda.alignment = Alignment(vertical="top", wrap_text=True)
    anchos = [16, 10, 9, 18, 20, 14, 14, 14, 14, 14, 12, 45, 25, 12, 20, 14, 14, 14, 12, 12]
    for i, ancho in enumerate(anchos, 1):
        det.column_dimensions[det.cell(row=1, column=i).column_letter].width = ancho
    ult = len(datos) + 1
    det.freeze_panes = "A2"
    det.auto_filter.ref = f"A1:T{ult}"

    # ---------------- hoja RESUMEN ----------------
    res = out.create_sheet("RESUMEN", 0)
    res.column_dimensions["A"].width = 42
    res.column_dimensions["B"].width = 12
    res.column_dimensions["C"].width = 10

    res["A1"] = "RESUMEN - CHECK LIST SS.HH DE LA CORPORACION"
    res["A1"].font = f_titulo
    res["A2"] = "Generado: " + datetime.now().strftime("%d/%m/%Y %H:%M")
    res["A2"].font = Font(name="Arial", size=9, italic=True, color="5F6368")

    res["A4"] = "TOTAL DE OBSERVACIONES"
    res["A4"].font = f_negrita
    res["B4"] = f"=COUNTA(DETALLE!A2:A{ult})"
    res["B4"].font = f_negrita

    def tabla(fila_ini, titulo, etiquetas, columna, con_blancos=False):
        """Escribe una tabla de conteo con fórmulas COUNTIF hacia DETALLE."""
        res.cell(row=fila_ini, column=1, value=titulo).font = f_sub
        fila = fila_ini + 1
        for texto, celda_v in (("Categoría", 1), ("Cantidad", 2), ("%", 3)):
            c = res.cell(row=fila, column=celda_v, value=texto)
            c.font = f_negrita
            c.fill = gris
        fila += 1
        for et in etiquetas:
            res.cell(row=fila, column=1, value=et).font = f_normal
            f = res.cell(row=fila, column=2,
                         value=f'=COUNTIF(DETALLE!${columna}$2:${columna}${ult},$A{fila})')
            f.font = f_normal
            p = res.cell(row=fila, column=3, value=f"=B{fila}/$B$4")
            p.font = f_normal
            p.number_format = "0.0%"
            fila += 1
        if con_blancos:
            res.cell(row=fila, column=1, value="SIN ESTATUS").font = f_normal
            res.cell(row=fila, column=2,
                     value=f"=COUNTBLANK(DETALLE!{columna}2:{columna}{ult})").font = f_normal
            p = res.cell(row=fila, column=3, value=f"=B{fila}/$B$4")
            p.font = f_normal
            p.number_format = "0.0%"
            fila += 1
        return fila + 1  # deja una fila en blanco

    def unicos(indice):
        vistos = []
        for vals in datos:
            v = str(vals[indice]).strip() if vals[indice] is not None else ""
            if v and v not in vistos:
                vistos.append(v)
        return sorted(vistos)

    estatus_datos = [e for e in unicos(16) if e not in ESTADOS]
    fila = tabla(6, "POR ESTATUS", ESTADOS + estatus_datos, "Q", con_blancos=True)
    fila = tabla(fila, "POR PRIORIDAD", ["ALTA", "MEDIA", "BAJA"], "T")
    fila = tabla(fila, "POR EDIFICIO", unicos(1), "B")
    fila = tabla(fila, "POR EMPRESA", unicos(4), "E")
    fila = tabla(fila, "POR UBICACION", unicos(3), "D")

    buffer = io.BytesIO()
    out.save(buffer)
    return buffer.getvalue()



# ---------------------------------------------------------------- PWA
# Íconos PNG incrustados (generados con la paleta del rediseño)
ICONO_192 = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAMAAAADACAIAAADdvvtQAAAIcUlEQVR42u3de0xV9wHA8d99gorhJQFBoLxFQEAGFyg+0tHWxtW5NFu1UZs2a+bWznZ12R9dtmn2x7p12dpsyWrTtZlUpizWaveyxkXlIXBHRQURr0plWF8XEUEe97k/MORyLtzClcc953y/8Q8859zrDefj7zw5aEJf3iyI/E3Lt4AARAAiABGAiABEACIAEYCIAEQAIgARgIgARAAiABGAiABEACIAEYAIQEQAIgARgAhARAAiABGACEBEACIAEYAIQEQAIgARgAhABCAiABGACEAEICIAEYAIQAQgIgARgCgw0/MtCMzu/rHC869hr2xhBCI/9Yw7JUDS8DtTA5xOgI9DjECy0ROY4xCAZKMnMA0BSE562IkmRekBEAGI4QdA6Jl8gXYkDyD0AIgtF4DQI8fhB0DoARB6AIQemeoBEHoARABi+JHp8AMg9AAIPQBCj0z1AAg9AEIPgNAjUz0AQg+A0AMgAhCpcfgBEHoAhB4AoUfwfCB1PilH5XpkMAIF8pNy0CMC/PlAgfywHPQE+ggUyA/LQU+gAwrkh+WgR1FHYYr5EU8AqcIQw48yzwPNjiH0KPlE4kwbQo+cAPn3fZ85Q+iR3wgUOIbQo65rYdNrCD0yBuT3apguQ+iR/Qg0h4bQo5BN2JwYQo+i9oHmfFtGst+Jnk1DDD/KPAqbHUPoUfJh/EwbQo/yzwPNnCH0qOVE4kwYQo+6zkRPryH0qPFSxnQZQo96r4U9vCH0CH5r8yyfLUSP0q7Gs0YBJBtDYFXs/UCzsGrRo/AbymZ0BaNHFXckztBqRo+Kbmmd9pWNHtU9H2gaVzl6VPqAqWlZ8ehR9RPKHnL1o4dH3PmPAD0AggKA5sgQ5gDkvwn0AMh/GegBkP8+0CPUeT+QUNazpwFEbMKIAEQAIgARgIgARAAiAJHC0svuE9f/9M2li+P8fvmwwxH92guFSalHd/zC95L5O3d0WG/5WGCjqezdLd/zscCuw1W//+xTIcSVX/8pYkHIV362t4/+feeh/fJaHbrgouXy+sQ2p2PIbo8JDZtnMEpmWfv79jfWNnd2jPxp+7Lr8u2bvYMDMaFhWs2Dsdbpcv3234ccTueg3XbfNpwUFa3RaMb9h3oHB2osbT4+yUuryvMSkrynu9zuGkvbwc8bjrScvtF7VwjhdLv7h4eiFobOMxq9lx+wDR9rO/fJ6cYjLc3Xeu4ILmXMQsEG44cvvvxUzgrPic2dHWt+83PvhdOiF1e89OrSmLjREWh0VnFy+sev/GS+Mai7vy8yZKHnqzq7rbk7X3e73RN9hpZfvr0kPHLIbgseS/kHH71XWV/tvXyQXr/7+e9vyC+SaCvY9WPfQx37QNPfkN1W2VA9yYUtN6+vf+dX13q6x9kgXrn4QfUxIcSF69favuzynJUQuag0NWOi9yxITFkSHllruXC7755k1sn28xNtPSvrT0omfmG9JV898t6J7hscnPzCt/p6t77/B6fL5T2rocMy8sW+xhrJrOdMKyd6wy2lq4UQFaeOe8+yOx0TverekPQz2xwOjsLkUdPVy+MOWqMQ95vrXGM3WN/MKxp3r2We0fhMQfG9wYFPTps5jFdRP9z7vucOkKQbvT0n2ls9p4QEB6/PLfReckO+aWHwvL/999SQ3QYghbfjyfWdb+2e5ML7Gmq8j9W9F9tauloIsafuOOeBlA/oWytMk1/48Bnz/eEhzymrM7JiwyI8p6RExZSkZJzrunrmf18ASOGA1ucVZsclTH75QZvtcPOY3RqtRvNs0aOeUzaXrGL4UTggvVYXH7Foe/m6957fNtXX/tVrK7bJYyum02o3mVYO2e1V5jr0yPJSho/yEpIe/oGb1Za2rp7uJeGRo1PSo2MLElOarl4WQjyelRsTGlZlrusdHECP0gBZ+/s+bTYLIQw6XVx4ZElKerDBONU3cbvdVea61594WjIIjQDaWrKG7ZdiAXXdsf5o34ejf10UsnDPd7eXpi6d+lasWgLoma8Vv3Fgb9j8BU9k53ZYb9VeugAd5e9EW/v7tlXstjudU32h5eb1kfFmtPD5IU/m5G00lem1uoq64z4ukAFIUXV2W/9xtsmPF3qfEHrOtHJzySqny1XpNQtASu7P1cc+OnVyqq860FRvG3tJa212fnp07NHWMzd6e3CjIkDVF8+/8fFeHwsYdHohhEE/Znfwzv3+o61nPKeM3Db0F3afuaVVUpDBIIQICQr6yhNCN+/1SlQRgERsWLgQIjEySnJr4mctzT0D/Z5TKutPOlxOvmMKATQycogxd1kE+fE+BY+kCCEWBAWXpGRI7p09+Hmj55Rx96UMOp3wulvS9+ZScmcIgOamNRlZwuvW1czYJVN6k9UZWRvyH1xtfWfTi2nRiz3nVplrR7+utVy4fPuG5OUlKRkxoeGSietyCyb650xJaZIpiZFRufGPCB7vMmttMq3MjotfkZgsGTBGh43/nD937lrnBzXHrt+d8HBpWWz8xqJHTcnppuQ0yWnohg7LifbWN/950O12azSas7t+Fx+xSAixbc+7+xofePp6Zk5ZWmZWXMJjmdl6rc77/c0dlxquXDzUbDZ3XBJCvFD2WGxYRObiuKdyVui0Wq87GJ0n2lvPdl090nK64YpFcCZ6Rnu1fJ2PH+sx6vRrc/LX5uSfaG/1Aag4OW17+bpx/j9pNMXJ6cXJ6W/965DD7XS73Qea6l97/BuSmw+fLSr7TmGpjw9ZmJRamJTaPzw8AuhnT3/bx4/1GHS68mXLy5ct12o0sgPEA6aIozACEAGIAEQEIAIQAYgARAQgAhABiABEBCACEAGIAEQEIAIQAYgARAAiAhABiABEACICEAGIAEQAIgIQAYgARAAiAhABiAKh/wNNDDBwt63lNgAAAABJRU5ErkJggg==")
ICONO_512 = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAIAAAB7GkOtAAAWYElEQVR42u3deXRVB53A8Zs9QEIIe1mbsIZASdihtMBUaW1rUaGLlOo4jto5OI56Ro+j1ekZZ9zO6UzVqrXqjJZait1sa8dWKgXKHvZQwr4vYQn7ngTmD5waEgi5yUuz3M/nLwh523283/e9e++7Ny5j+rQAgOiJtwgABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABABAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABAAAAQBAAAB4/yRaBEDdHXtixrX+qc3nH7J8Gqe4jOnTLAUg5nNfCQQAMPplILANAIj69K/jZREAoKlOfw0QACC6018DAtsAgGiO/sD2AJ8AgIhPf58DBACI6PTXAAEAojv9NUAAgOhOfwQAiPT0VxoBAEAAACt/EAAgCtNfcgQA8N4fAQBMfwQAMP0RAMD0RwAA0x8BAEx/BAAw/REAAAQAaORv/50cRgAAK38QAMD0RwCA5j39rf8RAMB7fwQAiMz09/ZfAADTHwEATH8EALDeHwEAmtX09/ZfAIAoMv0FALDqHwEATH8EADD9EQDAbj8IAGC3HwQAMP0RAMD0RwDA9Df9EQAw/U1/BABMf9MfAQBMfwQAvP03/REAMP1NfwEATH/TXwAA0x8BAEx/b/8FADD9TX8BAEx/018AANPf9BcAwPRHAIDA130RAMAu/wgAYPojAIDpjwAAvu6LAAB2+0EAwPQ3/REAMP1NfwQATH/THwEA09/0RwDA9Df9EQDA9EcAwNt/0x8BANPf9EcAwPQ3/REAMP0RAMD09/ZfAADTHwEATH8EADD9EQDA9EcAANMfAQACB3tAAABf+EIAANMfAQDT3/RHAMD0BwEA09/bfwQATH/THwEA09/0RwDA9Df9qReJFkEzmy9e4aa//xsIQEQny3u/49WO6U/14jKmT7MUmvGbSi97b//9NyCwDSCaM8VO6Ka/6Y8ARHemaIDpDwIQ3Zly7IkZMuDrviAA0Z0phpTdfkAAojtTNMD0BwGI7kzRANMfBCC6M0UDTH8QAMML0x8BIHpjRQNMfwQAg4zAwR4QADQAX/hCAIjMZNEA0x8BwFzD9EcAMN2wfBAAzDjs9oMAYNJZJqY/AoB5Z2mY/ggApp7lYPojAJh9loDpjwBgAnrspj8CQNN9bfsc4H8IAoDPAR6y6Y8AoAEerOmPABCd13lEGmD6IwAQxQbY4IEA4O1eFEek3X4QALzmo9gA0x8BwCs/ig0w/REAiGIDTH8EACMgig0w/REANMC6IE86AoBxEJkG2OUfBEADotgA0x8EQAOi2ADTHwRAA6LYANstQAA0IIqz1W4/IAAaEMUGmP4gABoQxTlr+oMAaEAUp63pDwKgAVGcuaY/CIAGRHHymv4gABpg/nrWQABMk8g0wBe+QAA0IIoNMP1BADQgig0w/UEANCCKDbDhAQRAA6LYALv9gABoQBQbYPqDAGhAFBtg+oMAEMUGmP4gAESxAaY/CABRbIDpDwKAdUEWOwgAkWmAL3yBABDFBpj+IABEsQGmPwgAUWyAgz2AABDFBtjtBwSAKDbA9AcBIIoNMP1BAIhiA0x/EACi2ADTHwQA64IsPYiZuIzp0yyFiLNXpemPTwCYZVhiCAAmGpYVAoC5BggAGmARWUQIAAachQMCgDFnsYAAYNhZICAAGHkWBQgABp+FAAKA8efhgwBgCHrgIAAYhR4yCAAGogcLAoCxCAgAGuAxggBgPnp0IACYkh4XCABmpUcEAoCJafqDAGBumv4gAJiepj8IAGao6Q8CQIQnqekPAkAU56npDwKAqQoCAJFpgFCBABDF8Wr6gwAQxSFr+oMAEMVRa/qDABDFgWv6gwAQxbFr+oMAEMXha/qDABDFEWz6gwAQxQaY/iAAWBcECACRaYDegAAQxQaY/iAARLEBpj8IAFFsgOkPAkAUG2D6QyhxGdOnWQo0oGNPzDD6wScAfBQw/UEA0ADTHwKrgAisDjL6QQCIfAnMfRAAAALbAAAQAAAEAAABAEAAABAAAAEAQAAAEAAABAAAAQBAAAAQAAAEAAABAEAAABAAAAQAAAEAQAAAEAAABACAWEq0CJqruV/9t7weWU3oDn/6f37y4ooll/+86tHHstp3rI9beXrR3C88+6v6exRP//0X7skbHvOrfXLum1974ZmKP2mXlr71ez99356drz7/9FPzZntZ+QQATdhHh4xMTUquv+sf0KW7hYwAQGOUntri7sFD6+nKM1um9erQyUJGAGhgfy4qXLtnZ2l5eVO886+vWT5/0/pdJYfr48qnjrylnu72qF594+LiYniF50ovbD6wf05R4Yod26r8U+nra1dsOrDv4qVL9fdEHD97pnDPzjcKV23Yv9drqvmJy5g+zVJoxlISEwd3v3Fsn5wJ/Qfe3Ccnvrbj6ZGXZy7cXFTNLyQmJCTGJ7RKSWmVktqmZau2rdI6tW5zY/sOY3vnpKWmBiG3AVR8wz4iq/fkoaOnDB+dnBCbTVYXL13K/eY/7T92NOZL+3tTpj08/vY6Xklpeflb69e8s7lo0ZaNhXt2ll+8WP3vt0pJzetx46jsvuP65Y7u1S8pIaEut15+8eKcDYVvb1i3YsfWTcX7j5455UUkADQHnTPafP5vPvTwhNsT40PPiAefevz1tStqcaPJCYlTR93yr5Puy2yZVosAvCerfccfTf30LX0HxGRRPPrKrMdn/yHmS3jx17+b06VbXa6hYPuWT/33E3uOlgS1XQf1mXEf+PLEe1KTkmpx8XV7d33ylz/eeqjYiyWwCohmpvj4sUdennnfzx677pvKGLpQXvbrhW+P/PbXVu/aXpfr2X744KQff+9ahdhx+GCoa/t4PawF6pbZro7T/+yFC1Of+q9aT/8gCI6eOfWDP/7+rsf/41xpaS2eqft+9pjpLwA0Z3OKCqt5o11PDp48ftcPv7Og2pVIQQ1W3XzuN08WbN9S9Z+eW7Yw1FX169xlaM9esX2MH75y789jZ05vPrA/1DXMXr/m0MkTdb8nK3Zu/fXCt8Ne6u2idfuOHfECEQCaudqtzKmj0+fPPfjU4xuL99XlSsoulj/66qyqPy/av6dwz86QHwLGxvYBfmzIyIp/nVWw8HxZuLfhq+r2IamiV1cXhL1IDG8dAaDx2lVyqEFu9/jZM/c/+dixM6frciULN29Yv2931Z/PXLYg1PVMHjYqVluVgyDo37nr8KzewZXfOAt7JYdj8fY/qNU6sSAISk6d8NIQAJq/0+fPN9RN7zh8sO5fxH3naquSXihYHGrbRmbLtDsG5cfqcT00ZlzFv67cue3dvbvDXkn5pYsx/LxVi/1/vDQEAOrXq6sLZiyeV5drWL59a3C1zQxzNhSGup6po2KzKTgpIeH+EVesUKrjAwQBoNn6xkvPHjhxvNYXP3yNlRXPLQ23FugDA27qkN667g/njkH57dPSgwo787ywfLFnGQGAqzhx9sw3XvptUPudHU9fa+P2yXNna349ifEJ9w4fU/eH89Do8RX/+vLKpaHuBggA0fLC8sUrd26r3WXPXGMbxrnS0ldWLQve38NC3NAm87acQUHdNv+CABAt33n9xZhfZ9h9gQZ27TGoW8+63OKDI29NiP/r62jTgX1Ltm3y5CIAUJ231q+N+SHGFm3ZGPYQclPr8IWAuLi4B0ffWvEnMxbZ/IsAQA3U4gur1bt06dKsgnDfCp4ybEwtDo502dg+ORVPXFNaXj5z6TueVgJnBIOR2X3e/PK3Lv/5C8/+qurK8Sfnvvnk3Ddje6Ozli34yh2Tav77HdJbfzB38B8LVwa12fx7xe7/fyxcefjUSc87PgFAkNe9Ac5JueVg8fIdW4P63xSc0aLlPXnDApt/EQAIrnaClAa53edCbgq+fVDedY9WXdW9w8dUPLvk3qMlc4oKPekIAASpSckTcwc3yE2/uHzJhfKyIMxJC6YMGxX2VqaNumLz7zNL5tfryblAAGgyPph7U6uU1Aa56aNnTs1+d00Q7rAQtwYh9x/N65FV8VDVzyye70lHAODyG+RxDXjrM0MeFiK/R1b/zl2D2h79be6GdbuPHPakIwAQ9Ol0wwcbaP3PZX9at/rI6XDns/14jY8Nl5KYeN+Vx5Cw+RcBgL945O4ptT4BfRCjU1G+FPLEZ/cPH1PxO73VuHvwsIobjUtOnfzftSs96QgABJPyR0zKH9HgdyPsN8I6Z2RO6D8wqNH6n/FXfvNgYahtziAANE+3D8z7+Scebgz3pGD7li0Hi2N+svjubduP6zsgsP4HAYD3pKWmPjrp/pmf+3JqUlIjuUuzQn4h4K6bhrZu0bL635k2+ta4Cmu3lm3fsqF4r2efwKEgiJQWyckZLVp1SG89pGf2qOw+dw8elp7aolHdw+eWLfz6XZPjarw1IjUp6aNDRv7m2kcoio+Le/DKHUa9/UcAaOZ++9kvNsW7vfvI4cVbN47p3T8Ic1iIagIwvv/Abpntggqn3g27qRkCq4AgaJRfCBiZ3adXh85BzY7+9uKKJWcunLeQEQBojH6/atm50guhLvLANc4Q0LZV2p03DQ2s/0EAoEk4ee7s6yF30v/4yLFX3Wxw3/CbUxL/uta0aN+esIcdhcA2AJqcR16euXBzUaUftkxJadOiVfv01kN6ZI3u3a9vpy6N887PWrZg8tAQx3rrltnulj458zetD6pd//Mbb/8RAKJg+6EDq3Ztv9a/Xt5qOiq771c+9JFKZ0hvDP5cVHjgxPFOrTOCMJuCKwVgSM/s3K7d3/vr+bKyWcsW+o9BYBUQBEGwZNumyT/5waOvzLrUyI6KXH7x4ovLF4e6yD35wysdyrTS2//X1y4/euaUJx0BgL96fPYfvv7Ss43tXs0M+Y2wlskpk/KHBxXObVBpJdLTC+d6rhEACKqe6bea9UUNonDPzvX7dge1PU/kR/KHV/yG8M6SQ/OqbCEAAYDg0qVL3/nDi43tXj0XcpX9zb3792jXPvjL+p/xFf/pmcXzLjn5FwIAVzV7/ZoN+xvXEXJ+V7Cw/OLFmv9+XFzcAyPGBkGQ3aHTmN79ggpbFJ5Z4uRfCABc2y/mz25U96f4+LF5G98Nwh8cdNrocRW/FvDnosL9x456fhEAuKbnly8+X1bWyNYChdsUnNW+49g+OVOv/GKwb/8iAHAdJ86eeXtDYaO6S6+tWX76/LlQF/nh1L/rnJH53l8Pnjz+RuEqTy4CANcxd8O6RnV/zl648OrqglAXqXRguJlLF5RdLPfMIgBwHY1tZ9BaHBy0khmL5npaCRwKAq5r9a7tE37wrct/3nXkcGO4S+9sLtp7tKRrhWP619zirRvDnmMSBICIOl9WVh8fAuLjKx+qMzE+vubfUfhdwaIvTfxwLW7X0d8IrAKCBn7bEp9Q6SdJCQlBve0LFPz/Nu3fr1xm4SMA0KABqDLuU5OSa37xjcX7avG55Pnli8OeWAYEAGKsVUpKpZ9kVDhQT40+BITfFGz3fwQAGl56SotKP8lslRbqGl5Ysbi0vDzUseTW7N5hySMA0MDapaVX+knXzLahrqHk1Mm31q/x9h8BgCamc0abSj/p2a5j2Ct5duk7NfzNc6UXflewyGJHAKDhZXfoVOknA7t2j7/aadyr8Wbh6hqe0uuV1QXHz56x2BEAaHj9b+gaVDmBV36P7FBXcqG87OWa7dZp/Q8CQDN61kO+WQ6CIDmxsXxnMD21xeDuN1b9ecUzONbQrBp8IWDboQOLtmys3V0Nu9CSEhpyITfsrSMAvE/apacHoVe7ZzaSOz8pf0TVL4IFQfDJmye0Drkz6NJtm3ccPlj978yow8m/umSE2zTdMb11rJZSm5atwl6kU5UtKwgAzVBul+5hLzK+X25juOepSclfmnj3Vf8po0XL70+ZFvpDQEF154ksu1j+bG1P/jU8q3daamqoi9zSd0CsFlRW+45N9ClGAKjf9T/TRo8Le6nbBgzK75HVsPc8KSHhZw99ttLxmYMrT+D1/XsfCnVYiOer3b3nT+vWHDhxvBZ3NSE+/mt3fizspcb2yRmZ3Scmy2rKsNFhLzKkZ/aE/gO9QCIlIXXETZZCdKQmJT12/99OzM0Ln434SfnDd5eUbChumDP93pYz6Nef/sdx13uXOuzGXlOGjT5fVrr5YHFp+fXPR3bk9KmJuXk3tLn6Cq5vvjyzFof/7Jie8cS0z9w5aEjYC8bFxU3KG7732JH1+/fUZVlNHjrqG3dPiQu/pefDg4cdOHH83b27nPA+IuIypk+zFIIIfHMq54ZuY/vkfGLMuC5t2tblqrYeKn5lVcGCzUWrdm6v4W6UQd22996WM2hibl5u13Crrc6VXnhr/do31q1+e8O6vUdLqvnNf5hwx3cnPxhc5RzCR3O/+cUankc+OSExq0PH3C49xvbNuW/YmLArf4IqRyt6bU3Bkq2b1u7eefBkTT+CdEhvPbRnr0+MGX/nTUPqcus7Dh98bc3yRVs2rt693dmPBYAm6V/u+tjI7L5d2mR2y2zXMjmlPm6i5NTJLQeLX1m97Kdz3ojh1Y7K7vvPd0zqf0PXbrU6av9VJ9qCzUWf/+0vg2t8rWz9v/+o6p5Rj7356rdfe776lWk//+TDXdu07dGuQ5c2bWuxb1UNF/KmA/teWrHkF/PfCq7c8/U/H/hUu7S0tq3S2rVK79S6TYvk5Jjf+slzZ7ccLN57tOSpebPnb1rvZRU4HwBNwu25eXn1vNa+XVp6u7T0klMnYxuA7I6dPjAglmsmb2zfsXWLlsE1AlB8/NjSbZtG9+oXXHnagBmL5wXXWy1277Ax78Ont9Fp/fYfP1opAC2Skx8YcXNQ/zvd5vfIyu+RNX/TegEIbASG5ue1NcuDKicOu+4eouATADR5P53zRmw/xEBgGwAAgVVAAAgAAAIAgAAAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAgAAAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAAAIAgAAAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAAAIAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAAAIAgAAAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAAAIAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAAAIAgAAAIAAACAAAAgCAAAAgAAAIAAACAIAAACAAAAgAAAIAgAAACAAAAgCAAAAgAAAIAAACAIAAACAAADR6/wex7nURiLERugAAAABJRU5ErkJggg==")

MANIFEST_JSON = json.dumps({
    "name": "CHECK LIST SS.HH FMI",
    "short_name": "CHECK LIST FMI",
    "description": "Registro y control de observaciones de SS.HH. de la corporación",
    "start_url": "/",
    "display": "standalone",
    "background_color": "#F6F5F1",
    "theme_color": "#0F6B5C",
    "icons": [
        {"src": "/icono-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
        {"src": "/icono-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
        {"src": "/icono-512.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
    ],
}, ensure_ascii=False)

SW_JS = """
self.addEventListener('install', e => self.skipWaiting());
self.addEventListener('activate', e => self.clients.claim());
self.addEventListener('fetch', e => {
  e.respondWith(fetch(e.request).catch(() =>
    new Response('<h3 style="font-family:sans-serif">Sin conexión a internet. Intenta de nuevo.</h3>',
                 {status: 503, headers: {'Content-Type': 'text/html; charset=utf-8'}})));
});
"""

# ---------------------------------------------------------------- HTML

ESTILO_BASE = """
* { box-sizing:border-box; margin:0; padding:0; }
:root {
  --tinta:#1F1E1B; --tinta2:#6E6B64; --tinta3:#8A877F;
  --fondo:#F6F5F1; --superficie:#FFFFFF;
  --linea:#E4E2DB; --linea2:#F0EEE8;
  --acento:#0F6B5C; --acento-osc:#0B5348; --acento-suave:#EAF2F0;
  --alerta:#B3261E; --ambar:#B45309;
}
body { font-family:"Segoe UI",system-ui,-apple-system,Arial,sans-serif; background:var(--fondo); padding:12px 16px; color:var(--tinta); }
a { color:var(--acento); text-decoration:none; }
.nav { margin:0 auto 14px; display:flex; gap:2px; align-items:center; flex-wrap:wrap; }
.nav .enlace { padding:7px 14px; font-size:13px; color:var(--tinta2); border-radius:8px; border:1px solid transparent; }
.nav .enlace:hover { color:var(--tinta); background:#EDEBE5; }
.nav .enlace.activo { color:var(--acento); font-weight:600; background:var(--superficie); border-color:var(--linea); }
"""

PAGINA_FORM = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#0F6B5C">
<meta name="apple-mobile-web-app-capable" content="yes">
<link rel="icon" type="image/png" href="/icono-192.png">
<link rel="apple-touch-icon" href="/icono-192.png">
<script>if ("serviceWorker" in navigator) navigator.serviceWorker.register("/sw.js");</script>
<title>CHECK LIST SS.HH DE LA CORPORACION</title>
<style>
__ESTILO__
.contenedor { max-width:700px; margin:0 auto; }
.cabecera { background:transparent !important; border:none !important; padding:14px 4px 20px !important; margin:0 !important; }
.cabecera::before { content:"FMI · MANTENIMIENTO"; display:block; font-size:11px; font-weight:600; letter-spacing:.14em; color:var(--acento); margin-bottom:10px; }
.cabecera h1 { font-family:Georgia,"Times New Roman",serif; font-size:30px; font-weight:400; color:var(--tinta); line-height:1.15; }
.cabecera p { font-size:13.5px; color:var(--tinta2); margin-top:10px; max-width:520px; }
form { background:var(--superficie); border:1px solid var(--linea); border-radius:14px; overflow:hidden; margin-bottom:40px; }
form .tarjeta { border:none; border-bottom:1px solid var(--linea2); margin:0; padding:22px 28px; background:transparent; }
.pregunta { font-size:12.5px; font-weight:600; letter-spacing:.05em; color:var(--tinta); margin-bottom:14px; }
.obligatorio { color:var(--alerta); }
div[id^="g-"] { display:flex; flex-wrap:wrap; gap:8px; }
label.opcion { display:inline-flex; align-items:center; gap:6px; padding:7px 15px; font-size:13px; color:#3E3C37; background:#FBFAF7; border:1px solid #D9D6CE; border-radius:999px; cursor:pointer; user-select:none; transition:border-color .12s, background .12s, color .12s; }
label.opcion:hover { border-color:var(--acento); color:var(--acento); }
label.opcion:has(input:checked) { background:var(--acento); border-color:var(--acento); color:#fff; }
label.opcion input { position:absolute; opacity:0; width:0; height:0; }
input[type=text] { width:100%; max-width:480px; border:none; border-bottom:1px solid #D9D6CE; font-size:14px; padding:7px 2px; outline:none; background:transparent; font-family:inherit; color:var(--tinta); }
input[type=text]:focus { border-bottom:1.5px solid var(--acento); }
.acciones { display:flex; justify-content:space-between; align-items:center; padding:20px 28px; background:var(--superficie); }
.btn { background:var(--acento); color:#fff; border:none; border-radius:9px; padding:11px 30px; font-size:14px; font-weight:600; cursor:pointer; }
.btn:hover { background:var(--acento-osc); }
.btn-borrar { background:none; border:none; color:var(--tinta2); font-size:13px; cursor:pointer; }
.btn-borrar:hover { color:var(--tinta); }
.mensaje { display:none; border-radius:10px; padding:13px 20px; margin-bottom:14px; font-size:13.5px; background:var(--superficie); }
.mensaje.ok { display:block; border:1px solid var(--linea); border-left:4px solid var(--acento); color:#0F5348; }
.mensaje.error { display:block; border:1px solid var(--linea); border-left:4px solid var(--alerta); color:var(--alerta); }
.grupo2col { display:flex; }
</style></head>
<body>
<div class="nav">
  <span class="enlace activo">📝 Registrar observación</span>
  <a class="enlace" href="/control">📊 Panel de control</a>
</div>
<div class="contenedor">
  <div class="tarjeta cabecera">
    <h1>CHECK LIST SS.HH DE LA CORPORACION</h1>
    <p>En este link se llenará las observaciones según necesidad</p>
  </div>
  <div id="mensaje" class="mensaje"></div>
  <form id="formulario">
    <div class="tarjeta">
      <div class="pregunta">PRIORIDAD <span class="obligatorio">*</span></div>
      <div id="g-prioridad"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">EDIFICIO <span class="obligatorio">*</span></div>
      <div id="g-edificio"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">PISO <span class="obligatorio">*</span></div>
      <div id="g-piso" class="grupo2col"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">EMPRESAS <span class="obligatorio">*</span></div>
      <div id="g-empresas"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">UBICACION <span class="obligatorio">*</span></div>
      <div id="g-ubicacion"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">SS.HH <span class="obligatorio">*</span></div>
      <div id="g-sshh"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">LAVATORIO</div>
      <div id="g-lavatorio" class="grupo2col"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">URINARIO</div>
      <div id="g-urinario" class="grupo2col"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">INODORO</div>
      <div id="g-inodoro" class="grupo2col"></div>
    </div>
    <div class="tarjeta">
      <div class="pregunta">PUERTAS DE INODOROS</div>
      <input type="text" name="puertas_inodoros" placeholder="Tu respuesta">
    </div>
    <div class="tarjeta">
      <div class="pregunta">DESCRIPCION</div>
      <input type="text" name="descripcion" placeholder="Tu respuesta">
    </div>
    <div class="tarjeta">
      <div class="pregunta">Comentario</div>
      <input type="text" name="comentario" placeholder="Tu respuesta">
    </div>
    <div class="tarjeta">
      <div class="pregunta">MESA DE LABATORIOS</div>
      <input type="text" name="mesa_lavatorios" placeholder="Tu respuesta">
    </div>
    <div class="acciones">
      <button type="submit" class="btn">Enviar</button>
      <button type="button" class="btn-borrar" onclick="borrar()">Borrar formulario</button>
    </div>
  </form>
</div>
<script>
const OPCIONES = __OPCIONES__;
const RADIOS = ["prioridad","edificio","piso","empresas","ubicacion"];
const CHECKS = ["sshh","lavatorio","urinario","inodoro"];
const EMOJI_PRIORIDAD = {"ALTA":"🔴","MEDIA":"🟡","BAJA":"🟢"};

for (const campo of [...RADIOS, ...CHECKS]) {
  const tipo = RADIOS.includes(campo) ? "radio" : "checkbox";
  const cont = document.getElementById("g-" + campo);
  cont.innerHTML = OPCIONES[campo].map(op => {
    const emoji = campo === "prioridad" ? (EMOJI_PRIORIDAD[op] || "") + " " : "";
    return `<label class="opcion"><input type="${tipo}" name="${campo}" value="${op}"> ${emoji}${op}</label>`;
  }).join("");
}

const form = document.getElementById("formulario");
const msg = document.getElementById("mensaje");
function borrar(){ form.reset(); msg.className = "mensaje"; }
function valorRadio(n){ const el = form.querySelector(`input[name="${n}"]:checked`); return el ? el.value : ""; }
function valoresCheck(n){ return [...form.querySelectorAll(`input[name="${n}"]:checked`)].map(c=>c.value).join(", "); }
function mostrar(t, err){ msg.textContent = t; msg.className = "mensaje " + (err ? "error" : "ok"); window.scrollTo({top:0, behavior:"smooth"}); }

form.addEventListener("submit", async e => {
  e.preventDefault();
  for (const [campo, nombre] of [["prioridad","PRIORIDAD"],["edificio","EDIFICIO"],["piso","PISO"],["empresas","EMPRESAS"],["ubicacion","UBICACION"]]) {
    if (!valorRadio(campo)) { mostrar("⚠️ La pregunta " + nombre + " es obligatoria.", true); return; }
  }
  if (!valoresCheck("sshh")) { mostrar("⚠️ La pregunta SS.HH es obligatoria.", true); return; }
  const datos = {
    prioridad: valorRadio("prioridad"),
    edificio: valorRadio("edificio"), piso: valorRadio("piso"),
    empresas: valorRadio("empresas"), ubicacion: valorRadio("ubicacion"),
    sshh: valoresCheck("sshh"), lavatorio: valoresCheck("lavatorio"),
    urinario: valoresCheck("urinario"), inodoro: valoresCheck("inodoro"),
    puertas_inodoros: form.puertas_inodoros.value.trim(),
    descripcion: form.descripcion.value.trim(),
    comentario: form.comentario.value.trim(),
    mesa_lavatorios: form.mesa_lavatorios.value.trim()
  };
  try {
    const r = await fetch("/guardar", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify(datos)});
    const res = await r.json();
    if (res.ok) { mostrar("✅ Observación registrada correctamente (fila " + res.fila + " del Excel).", false); form.reset(); }
    else mostrar("⚠️ Error al guardar: " + res.error, true);
  } catch { mostrar("⚠️ No se pudo conectar con el servidor.", true); }
});
</script>
</body></html>"""

PAGINA_CONTROL = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#0F6B5C">
<meta name="apple-mobile-web-app-capable" content="yes">
<link rel="icon" type="image/png" href="/icono-192.png">
<link rel="apple-touch-icon" href="/icono-192.png">
<script>if ("serviceWorker" in navigator) navigator.serviceWorker.register("/sw.js");</script>
<title>Panel de Control - CHECK LIST SS.HH</title>
<style>
__ESTILO__
.contenedor { width:100%; margin:0 auto; }
.tarjeta { background:var(--superficie); border:1px solid var(--linea); border-radius:12px; margin-bottom:10px; padding:12px 18px; }
.cab { display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap; }
h1 { font-family:Georgia,"Times New Roman",serif; font-size:19px; font-weight:400; color:var(--tinta); }
.btn-admin { border:1px solid var(--linea); border-radius:9px; padding:7px 15px; font-size:12.5px; font-weight:600; cursor:pointer; background:var(--superficie); color:var(--tinta2); }
.btn-admin:hover { color:var(--tinta); border-color:#C9C6BE; }
.btn-admin.activo { background:var(--tinta); border-color:var(--tinta); color:#fff; }
.btn-desc { display:inline-block; border-radius:9px; padding:7px 15px; font-size:12.5px; font-weight:600; border:1px solid var(--acento); color:var(--acento) !important; background:var(--superficie); }
.btn-desc:hover { background:var(--acento); color:#fff !important; }
.alerta { background:var(--superficie); border:1px solid var(--linea); border-left:4px solid var(--ambar); color:#7C4A03; border-radius:10px; padding:10px 16px; margin-bottom:10px; font-size:13px; font-weight:600; cursor:pointer; }
.alerta:hover { border-color:var(--ambar); }
.alerta.activa { background:var(--tinta); border-color:var(--tinta); color:#fff; }
.linea2 { display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:10px; }
.resumen { display:flex; gap:6px; flex-wrap:wrap; align-items:center; }
.chip { display:inline-flex; align-items:center; gap:6px; border-radius:999px; padding:4px 12px; font-size:12px; color:var(--tinta2); cursor:pointer; border:1px solid var(--linea); background:var(--superficie); white-space:nowrap; }
.chip:hover { border-color:#C9C6BE; color:var(--tinta); }
.chip.sel { border-color:var(--tinta); color:var(--tinta); font-weight:600; }
.chip i, .est i { width:8px; height:8px; border-radius:50%; display:inline-block; flex-shrink:0; }
.buscador { flex:1; min-width:200px; border:1px solid var(--linea); border-radius:9px; padding:7px 14px; font-size:13px; outline:none; background:var(--superficie); color:var(--tinta); }
.buscador:focus { border-color:var(--acento); }
table { width:100%; border-collapse:collapse; font-size:12px; }
th { background:var(--superficie); color:var(--tinta3); font-size:10px; font-weight:600; text-transform:uppercase; letter-spacing:.08em; padding:9px 6px; text-align:left; position:sticky; top:0; white-space:nowrap; z-index:2; border-bottom:1.5px solid #D9D6CE; }
td { border-bottom:1px solid var(--linea2); padding:6px; vertical-align:top; color:var(--tinta); line-height:1.4; }
tr:hover td { background:#FAF9F5; }
.desc { min-width:200px; max-width:340px; }
td input, td select { width:100%; border:1px solid var(--linea); border-radius:7px; padding:4px 6px; font-size:12px; font-family:inherit; outline:none; background:var(--superficie); color:var(--tinta); }
td input:focus, td select:focus { border-color:var(--acento); }
td input:disabled, td select:disabled { background:#F5F4F0; color:var(--tinta3); cursor:not-allowed; border-color:var(--linea2); }
.btn-g, .btn-e { border-radius:7px; padding:5px 8px; font-size:12px; cursor:pointer; display:block; margin-bottom:3px; width:100%; }
.btn-g { background:var(--acento); color:#fff; border:1px solid var(--acento); }
.btn-g:hover:not(:disabled) { background:var(--acento-osc); }
.btn-e { background:var(--superficie); color:var(--tinta2); border:1px solid var(--linea); }
.btn-e:hover:not(:disabled) { color:var(--tinta); border-color:#C9C6BE; }
.btn-g:disabled, .btn-e:disabled { background:#F5F4F0; color:#C0BDB5; border-color:var(--linea2); cursor:not-allowed; }
.est { display:inline-flex; align-items:center; gap:6px; font-size:11px; font-weight:600; white-space:nowrap; color:var(--tinta2); }
.tabla-scroll { overflow:auto; max-height:calc(100vh - 155px); padding:0; }
.tabla-scroll table { min-width:1500px; }
@media (max-width:900px) {
  table { font-size:11px; }
  h1 { font-size:16px; }
  .tabla-scroll { max-height:calc(100vh - 195px); }
}
.aviso { position:fixed; bottom:18px; right:18px; background:var(--tinta); color:#fff; border-radius:10px; padding:12px 18px; font-size:13px; display:none; z-index:99; }
/* --------- modal de edición --------- */
.fondo-modal { display:none; position:fixed; inset:0; background:rgba(31,30,27,.45); z-index:50; overflow-y:auto; padding:24px 12px; }
.fondo-modal.abierto { display:block; }
.modal { background:var(--superficie); border-radius:14px; max-width:760px; margin:0 auto; padding:26px 28px; border:1px solid var(--linea); }
.modal h2 { font-family:Georgia,"Times New Roman",serif; font-size:19px; font-weight:400; color:var(--tinta); margin-bottom:4px; }
.modal .sub { font-size:12.5px; color:var(--tinta2); margin-bottom:16px; }
.modal .grupo { margin-bottom:14px; }
.modal .etiqueta { font-size:10.5px; font-weight:600; color:var(--tinta3); text-transform:uppercase; letter-spacing:.08em; margin-bottom:6px; }
.modal select, .modal input[type=text] { width:100%; border:1px solid var(--linea); border-radius:8px; padding:8px 10px; font-size:13.5px; font-family:inherit; outline:none; background:var(--superficie); color:var(--tinta); }
.modal select:focus, .modal input[type=text]:focus { border-color:var(--acento); }
.modal .rejilla2 { display:grid; grid-template-columns:1fr 1fr; gap:0 18px; }
.modal .cajas { display:grid; grid-template-columns:repeat(auto-fill,minmax(140px,1fr)); }
.modal label.caja { display:flex; align-items:center; gap:7px; font-size:13px; padding:4px 0; cursor:pointer; color:var(--tinta); }
.modal label.caja input { accent-color:var(--acento); width:16px; height:16px; }
.modal .pie { display:flex; justify-content:flex-end; gap:10px; margin-top:18px; }
.modal .pie button { border:none; border-radius:9px; padding:10px 24px; font-size:14px; font-weight:600; cursor:pointer; }
.b-guardar { background:var(--acento); color:#fff; }
.b-guardar:hover { background:var(--acento-osc); }
.b-cancelar { background:#EDEBE5; color:var(--tinta); }
</style></head>
<body>
<div class="nav">
  <a class="enlace" href="/">📝 Registrar observación</a>
  <span class="enlace activo">📊 Panel de control</span>
  <a class="enlace" href="/dashboard">📈 Dashboard</a>
</div>
<div class="contenedor">
  <div class="tarjeta">
    <div class="cab">
      <h1>Panel de control de observaciones — cotizaciones y estatus</h1>
      <div style="display:flex; gap:8px; flex-wrap:wrap">
        <a class="btn-desc" href="/descargar">⬇️ Descargar resumen Excel</a>
        <a class="btn-admin" style="text-decoration:none; display:inline-block" href="/historial">🕘 Historial</a>
        <button id="btn-admin" class="btn-admin" onclick="alternarAdmin()">🔒 Modo administrador</button>
      </div>
    </div>
    <div class="linea2">
      <input id="buscar" class="buscador" type="text" placeholder="🔎 Buscar por empresa, ubicación, descripción, proveedor, PO...">
      <div class="resumen" id="resumen"></div>
    </div>
  </div>
  <div id="alerta-viejos" class="alerta" style="display:none" onclick="alternarViejos()"></div>
  <div class="tarjeta tabla-scroll" style="padding:0">
    <table>
      <thead><tr>
        <th>#</th><th>Fecha</th><th>Prioridad</th><th>Edificio</th><th>Piso</th><th>Ubicación</th><th>Empresa</th>
        <th>SS.HH</th>
        <th>Lavatorio</th>
        <th>Mesa de Labatorios</th>
        <th>Inodoro</th>
        <th>Puertas de Inodoros</th>
        <th>Urinario</th>
        <th class="desc">Descripción</th>
        <th class="desc">Comentario</th>
        <th>Dispensador</th>
        <th style="min-width:150px">COMENTARIO FMI / PROVEEDOR</th>
        <th style="min-width:120px">COTIZACION PROVEEDOR</th>
        <th style="min-width:140px">ESTATUS</th>
        <th style="min-width:120px">COTIZACION PO</th>
        <th style="min-width:90px">MONTO (S/)</th>
        <th style="min-width:60px">Acción</th>
      </tr></thead>
      <tbody id="cuerpo"></tbody>
    </table>
  </div>
</div>

<!-- Modal de edición completa (solo administrador) -->
<div class="fondo-modal" id="fondo-modal">
  <div class="modal">
    <h2>✏️ Editar observación — fila <span id="m-fila"></span></h2>
    <div class="sub" id="m-fecha"></div>
    <div class="rejilla2">
      <div class="grupo"><div class="etiqueta">Prioridad</div><select id="m-prioridad"></select></div>
      <div class="grupo"><div class="etiqueta">Edificio</div><select id="m-edificio"></select></div>
      <div class="grupo"><div class="etiqueta">Piso</div><select id="m-piso"></select></div>
      <div class="grupo"><div class="etiqueta">Empresa</div><select id="m-empresas"></select></div>
      <div class="grupo"><div class="etiqueta">Ubicación</div><select id="m-ubicacion"></select></div>
    </div>
    <div class="grupo"><div class="etiqueta">SS.HH</div><div class="cajas" id="m-sshh"></div></div>
    <div class="grupo"><div class="etiqueta">Lavatorio</div><div class="cajas" id="m-lavatorio"></div></div>
    <div class="grupo"><div class="etiqueta">Urinario</div><div class="cajas" id="m-urinario"></div></div>
    <div class="grupo"><div class="etiqueta">Inodoro</div><div class="cajas" id="m-inodoro"></div></div>
    <div class="rejilla2">
      <div class="grupo"><div class="etiqueta">Puertas de inodoros</div><input type="text" id="m-puertas"></div>
      <div class="grupo"><div class="etiqueta">Mesa de labatorios</div><input type="text" id="m-mesa"></div>
    </div>
    <div class="grupo"><div class="etiqueta">Descripción</div><input type="text" id="m-descripcion"></div>
    <div class="grupo"><div class="etiqueta">Comentario</div><input type="text" id="m-comentario"></div>
    <div class="grupo"><div class="etiqueta">Dispensador</div><input type="text" id="m-dispensador"></div>
    <div class="pie">
      <button class="b-cancelar" onclick="cerrarModal()">Cancelar</button>
      <button class="b-guardar" id="m-guardar" onclick="guardarEdicion()">Guardar cambios</button>
    </div>
  </div>
</div>

<div class="aviso" id="aviso"></div>
<script>
const DATOS = __DATOS__;
const ESTADOS = __ESTADOS__;
const OPCIONES = __OPCIONES__;
const COLORES = {
  "PENDIENTE":       ["#fce8e6","#a50e0e"],
  "SOLI.COTI.PROV":  ["#fef7e0","#8a5a00"],
  "COTIZACION":      ["#fff0d4","#a05a00"],
  "APROBADO":        ["#e8f0fe","#174ea6"],
  "EN_ EJECUCION":   ["#f3e8fd","#5b2b91"],
  "ATENDIDO":        ["#e6f4ea","#1e4620"],
  "":                ["#f1f3f4","#5f6368"]
};
const COLORES_PRIORIDAD = {
  "ALTA":  ["#fce8e6","#a50e0e"],
  "MEDIA": ["#fef7e0","#8a5a00"],
  "BAJA":  ["#e6f4ea","#1e4620"],
  "":      ["#f1f3f4","#5f6368"]
};
let filtroEstado = null;
let claveAdmin = sessionStorage.getItem("claveAdmin") || null;
let rolAdmin = sessionStorage.getItem("rolAdmin") || null;
let filaEnEdicion = null;

function esAdmin(){ return claveAdmin !== null; }        /* cualquier clave válida */
function esAdminTotal(){ return esAdmin() && rolAdmin === "admin"; }

function salirAdmin(){
  claveAdmin = null; rolAdmin = null;
  sessionStorage.removeItem("claveAdmin");
  sessionStorage.removeItem("rolAdmin");
}

async function alternarAdmin(){
  if (esAdmin()) {
    salirAdmin();
    aviso("🔒 Modo edición desactivado");
    refrescarAdmin(); pintar();
    return;
  }
  const clave = prompt("Ingresa la clave (administrador o edición limitada):");
  if (!clave) return;
  try {
    const r = await fetch("/verificar", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify({clave})});
    const res = await r.json();
    if (res.ok) {
      claveAdmin = clave; rolAdmin = res.rol;
      sessionStorage.setItem("claveAdmin", clave);
      sessionStorage.setItem("rolAdmin", res.rol);
      aviso(res.rol === "admin" ? "🔓 Modo administrador: puedes editar todo"
                                : "🔓 Modo limitado: solo COMENTARIO FMI / PROVEEDOR y COTIZACION PROVEEDOR");
      refrescarAdmin(); pintar();
    } else aviso("⚠️ Clave incorrecta");
  } catch { aviso("⚠️ No se pudo conectar con el servidor"); }
}

function refrescarAdmin(){
  const b = document.getElementById("btn-admin");
  if (esAdminTotal())      { b.textContent = "🔓 ADMINISTRADOR activo (clic para salir)"; b.className = "btn-admin activo"; }
  else if (esAdmin())      { b.textContent = "🔓 Edición limitada activa (clic para salir)"; b.className = "btn-admin activo"; }
  else                     { b.textContent = "🔒 Modo administrador"; b.className = "btn-admin"; }
}

function esc(s){ const d=document.createElement("div"); d.textContent=s; return d.innerHTML; }
/* compacta "LAVATORIO 1, LAVATORIO 2" -> "1, 2" (el valor completo se conserva en el Excel) */
function compacto(valor, palabra){
  if (!valor) return "";
  return valor.split(",").map(s => s.trim().replace(new RegExp("^" + palabra + "\\\\s*", "i"), "")).filter(Boolean).join(", ");
}

function pintarResumen(){
  const cont = document.getElementById("resumen");
  const conteo = {};
  for (const r of DATOS) { const e = r.estatus || "SIN ESTATUS"; conteo[e] = (conteo[e]||0)+1; }
  const claves = ["PENDIENTE","SOLI.COTI.PROV","COTIZACION","APROBADO","EN_ EJECUCION","ATENDIDO","SIN ESTATUS"];
  cont.innerHTML = `<span class="chip ${filtroEstado===null?"sel":""}" onclick="filtrarEstado(null)"><i style="background:#54524C"></i>TODOS: ${DATOS.length}</span>` +
    claves.filter(k => conteo[k]).map(k => {
      const c = COLORES[k==="SIN ESTATUS" ? "" : k];
      return `<span class="chip ${filtroEstado===k?"sel":""}" onclick="filtrarEstado('${k}')"><i style="background:${c[1]}"></i>${k}: ${conteo[k]}</span>`;
    }).join("");
}
function filtrarEstado(e){ filtroEstado = (filtroEstado===e) ? null : e; pintar(); }

/* --------- alerta de casos antiguos abiertos --------- */
let filtroViejos = false;
function esViejo(r){ return r.estatus !== "ATENDIDO" && r.dias !== null && r.dias > 180; }
function alternarViejos(){ filtroViejos = !filtroViejos; pintar(); }
function pintarAlerta(){
  const a = document.getElementById("alerta-viejos");
  const v6 = DATOS.filter(esViejo).length;
  const v3 = DATOS.filter(r => r.estatus !== "ATENDIDO" && r.dias !== null && r.dias > 90 && r.dias <= 180).length;
  if (v6 === 0 && v3 === 0) { a.style.display = "none"; filtroViejos = false; return; }
  a.style.display = "block";
  a.className = "alerta" + (filtroViejos ? " activa" : "");
  if (filtroViejos) {
    a.innerHTML = "⚠️ Mostrando los " + v6 + " casos con más de 6 meses abiertos, del más antiguo al más reciente — clic para quitar el filtro";
  } else {
    a.innerHTML = "⚠️ " + v6 + " casos llevan más de 6 meses abiertos" +
      (v3 ? " &nbsp;·&nbsp; " + v3 + " casos entre 3 y 6 meses" : "") +
      " &nbsp;—&nbsp; clic para ver los más antiguos";
  }
}

function pintar(){
  pintarResumen();
  pintarAlerta();
  const disLim = esAdmin() ? "" : "disabled";        /* columnas FMI/Proveedor: ambas claves */
  const disTot = esAdminTotal() ? "" : "disabled";   /* resto: solo administrador */
  const q = document.getElementById("buscar").value.toLowerCase();
  const cuerpo = document.getElementById("cuerpo");
  const visibles = DATOS.filter(r => {
    if (filtroViejos && !esViejo(r)) return false;
    const est = r.estatus || "SIN ESTATUS";
    if (filtroEstado && est !== filtroEstado) return false;
    if (!q) return true;
    return Object.values(r).join(" ").toLowerCase().includes(q);
  }).sort((a,b) => filtroViejos ? (b.dias - a.dias) : (b.fila - a.fila));
  cuerpo.innerHTML = visibles.map(r => {
    const c = COLORES[r.estatus] || COLORES[""];
    const opts = [""].concat(ESTADOS).map(e =>
      `<option value="${esc(e)}" ${e===r.estatus?"selected":""}>${e || "(sin estatus)"}</option>`).join("");
    const pc = COLORES_PRIORIDAD[r.prioridad] || COLORES_PRIORIDAD[""];
    return `<tr id="fila-${r.fila}">
      <td>${r.fila}</td><td>${esc(r.fecha)}</td>
      <td><span class="est"><i style="background:${pc[1]}"></i>${r.prioridad || "-"}</span></td>
      <td>${esc(r.edificio)}</td><td>${esc(r.piso)}</td>
      <td>${esc(r.ubicacion)}</td><td>${esc(r.empresas)}</td><td>${esc(compacto(r.sshh, "SS-HH"))}</td>
      <td>${esc(compacto(r.lavatorio, "LAVATORIO"))}</td>
      <td>${esc(r.mesa)}</td>
      <td>${esc(compacto(r.inodoro, "INODORO"))}</td>
      <td>${esc(r.puertas)}</td>
      <td>${esc(compacto(r.urinario, "URINARIO"))}</td>
      <td class="desc">${esc(r.descripcion)}</td>
      <td class="desc">${esc(r.comentario)}</td>
      <td>${esc(r.dispensador)}</td>
      <td><input ${disLim} data-f="${r.fila}" data-c="comentario_mili" value="${esc(r.comentario_mili)}"></td>
      <td><input ${disLim} data-f="${r.fila}" data-c="proveedor" value="${esc(r.proveedor)}"></td>
      <td>
        <span class="est"><i style="background:${c[1]}"></i>${r.estatus || "SIN ESTATUS"}</span><br>
        <select ${disTot} data-f="${r.fila}" data-c="estatus" style="margin-top:4px">${opts}</select>
      </td>
      <td><input ${disTot} data-f="${r.fila}" data-c="po" value="${esc(r.po)}"></td>
      <td><input ${disTot} data-f="${r.fila}" data-c="monto" value="${esc(r.monto)}" placeholder="0.00"></td>
      <td>
        <button class="btn-g" ${disLim} title="Guardar seguimiento" onclick="guardar(${r.fila}, this)">💾</button>
        <button class="btn-e" ${disTot} title="Editar toda la observación (solo administrador)" onclick="abrirModal(${r.fila})">✏️</button>
      </td>
    </tr>`;
  }).join("");
}

async function guardar(fila, btn){
  if (!esAdmin()) { aviso("🔒 Activa el modo administrador para editar"); return; }
  const reg = DATOS.find(r => r.fila === fila);
  const datos = { fila, clave: claveAdmin };
  /* solo se envían los campos habilitados según la clave usada */
  document.querySelectorAll(`#fila-${fila} [data-c]:not(:disabled)`).forEach(el => { datos[el.dataset.c] = el.value; });
  btn.disabled = true;
  try {
    const r = await fetch("/actualizar", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify(datos)});
    const res = await r.json();
    if (res.ok) {
      for (const campo of ["comentario_mili","proveedor","estatus","po","monto"]) {
        if (campo in datos) reg[campo] = datos[campo];
      }
      aviso("✅ Fila " + fila + " actualizada en el Excel");
      pintar();
    } else aviso("⚠️ " + res.error);
  } catch { aviso("⚠️ No se pudo conectar con el servidor"); }
  btn.disabled = false;
}

/* --------- edición completa (modal) --------- */
function llenarSelect(id, opciones, valor){
  const sel = document.getElementById(id);
  const lista = opciones.includes(valor) || valor === "" ? opciones : [valor, ...opciones];
  sel.innerHTML = `<option value=""></option>` + lista.map(o =>
    `<option value="${esc(o)}" ${o===valor?"selected":""}>${o}</option>`).join("");
}
function llenarCajas(id, opciones, valorActual){
  const marcados = valorActual.split(",").map(s => s.trim()).filter(Boolean);
  const extras = marcados.filter(m => !opciones.includes(m));
  document.getElementById(id).innerHTML = [...opciones, ...extras].map(o =>
    `<label class="caja"><input type="checkbox" value="${esc(o)}" ${marcados.includes(o)?"checked":""}> ${o}</label>`).join("");
}
function valoresCajas(id){
  return [...document.querySelectorAll(`#${id} input:checked`)].map(c => c.value).join(", ");
}

function abrirModal(fila){
  if (!esAdminTotal()) { aviso("🔒 Solo el administrador puede editar la observación completa"); return; }
  const r = DATOS.find(x => x.fila === fila);
  filaEnEdicion = fila;
  document.getElementById("m-fila").textContent = fila;
  document.getElementById("m-fecha").textContent = "Registrada: " + r.fecha;
  llenarSelect("m-prioridad", OPCIONES.prioridad, r.prioridad);
  llenarSelect("m-edificio", OPCIONES.edificio, r.edificio);
  llenarSelect("m-piso", OPCIONES.piso, r.piso);
  llenarSelect("m-empresas", OPCIONES.empresas, r.empresas);
  llenarSelect("m-ubicacion", OPCIONES.ubicacion, r.ubicacion);
  llenarCajas("m-sshh", OPCIONES.sshh, r.sshh);
  llenarCajas("m-lavatorio", OPCIONES.lavatorio, r.lavatorio);
  llenarCajas("m-urinario", OPCIONES.urinario, r.urinario);
  llenarCajas("m-inodoro", OPCIONES.inodoro, r.inodoro);
  document.getElementById("m-puertas").value = r.puertas;
  document.getElementById("m-mesa").value = r.mesa;
  document.getElementById("m-descripcion").value = r.descripcion;
  document.getElementById("m-comentario").value = r.comentario;
  document.getElementById("m-dispensador").value = r.dispensador;
  document.getElementById("fondo-modal").classList.add("abierto");
}
function cerrarModal(){ document.getElementById("fondo-modal").classList.remove("abierto"); filaEnEdicion = null; }

async function guardarEdicion(){
  if (filaEnEdicion === null) return;
  const btn = document.getElementById("m-guardar");
  const datos = {
    fila: filaEnEdicion, clave: claveAdmin,
    prioridad: document.getElementById("m-prioridad").value,
    edificio: document.getElementById("m-edificio").value,
    piso: document.getElementById("m-piso").value,
    empresas: document.getElementById("m-empresas").value,
    ubicacion: document.getElementById("m-ubicacion").value,
    sshh: valoresCajas("m-sshh"),
    lavatorio: valoresCajas("m-lavatorio"),
    urinario: valoresCajas("m-urinario"),
    inodoro: valoresCajas("m-inodoro"),
    puertas: document.getElementById("m-puertas").value.trim(),
    mesa: document.getElementById("m-mesa").value.trim(),
    descripcion: document.getElementById("m-descripcion").value.trim(),
    comentario: document.getElementById("m-comentario").value.trim(),
    dispensador: document.getElementById("m-dispensador").value.trim()
  };
  btn.disabled = true;
  try {
    const r = await fetch("/editar", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify(datos)});
    const res = await r.json();
    if (res.ok) {
      const reg = DATOS.find(x => x.fila === filaEnEdicion);
      Object.assign(reg, datos);
      aviso("✅ Observación de la fila " + filaEnEdicion + " editada en el Excel");
      cerrarModal();
      pintar();
    } else aviso("⚠️ " + res.error);
  } catch { aviso("⚠️ No se pudo conectar con el servidor"); }
  btn.disabled = false;
}

document.getElementById("fondo-modal").addEventListener("click", e => { if (e.target.id === "fondo-modal") cerrarModal(); });
function aviso(t){
  const a = document.getElementById("aviso");
  a.textContent = t; a.style.display = "block";
  clearTimeout(a._t); a._t = setTimeout(() => a.style.display = "none", 3500);
}
document.getElementById("buscar").addEventListener("input", pintar);

/* si había sesión de admin guardada, revalidarla */
if (claveAdmin) {
  fetch("/verificar", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify({clave:claveAdmin})})
    .then(r => r.json()).then(res => {
      if (!res.ok) salirAdmin();
      else { rolAdmin = res.rol; sessionStorage.setItem("rolAdmin", res.rol); }
      refrescarAdmin(); pintar();
    })
    .catch(() => { refrescarAdmin(); pintar(); });
} else { refrescarAdmin(); pintar(); }
</script>
</body></html>"""

PAGINA_DASHBOARD = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#0F6B5C">
<meta name="apple-mobile-web-app-capable" content="yes">
<link rel="icon" type="image/png" href="/icono-192.png">
<link rel="apple-touch-icon" href="/icono-192.png">
<script>if ("serviceWorker" in navigator) navigator.serviceWorker.register("/sw.js");</script>
<title>Dashboard - CHECK LIST SS.HH</title>
<style>
__ESTILO__
.contenedor { width:100%; max-width:1500px; margin:0 auto; }
.tarjeta { background:var(--superficie); border:1px solid var(--linea); border-radius:12px; margin-bottom:10px; padding:16px 20px; }
h1 { font-family:Georgia,"Times New Roman",serif; font-size:19px; font-weight:400; color:var(--tinta); }
h2 { font-size:13px; font-weight:600; color:var(--tinta); margin-bottom:2px; letter-spacing:.02em; }
.sub { font-size:11.5px; color:var(--tinta2); margin-bottom:12px; }
.kpis { display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:10px; margin-bottom:10px; }
.kpi { background:var(--superficie); border:1px solid var(--linea); border-radius:12px; padding:14px 18px; }
.kpi .valor { font-family:Georgia,"Times New Roman",serif; font-size:28px; font-weight:400; color:var(--tinta); line-height:1.1; }
.kpi .nombre { font-size:10.5px; font-weight:600; text-transform:uppercase; letter-spacing:.08em; color:var(--tinta3); margin-top:5px; }
.rejilla { display:grid; grid-template-columns:1fr 1fr; gap:10px; }
@media (max-width:1000px){ .rejilla { grid-template-columns:1fr; } }
.leyenda { display:flex; gap:14px; font-size:11.5px; color:var(--tinta2); margin-bottom:6px; }
.leyenda span::before { content:""; display:inline-block; width:9px; height:9px; border-radius:50%; margin-right:5px; vertical-align:-1px; }
.leyenda .l-abierto::before { background:#C2410C; }
.leyenda .l-atendido::before { background:#0F6B5C; }
svg text { font-family:"Segoe UI",system-ui,Arial,sans-serif; }
.aviso-vacio { color:var(--tinta2); font-size:12.5px; padding:20px 0; text-align:center; }
</style></head>
<body>
<div class="nav">
  <a class="enlace" href="/">📝 Registrar observación</a>
  <a class="enlace" href="/control">📊 Panel de control</a>
  <span class="enlace activo">📈 Dashboard</span>
  <a class="enlace" href="/historial">🕘 Historial</a>
</div>
<div class="contenedor">
  <div class="tarjeta" style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px">
    <h1>📈 Dashboard de observaciones</h1>
    <span style="font-size:11.5px;color:#5f6368">Datos al día — se actualiza al recargar la página</span>
  </div>
  <div class="kpis" id="kpis"></div>
  <div class="rejilla">
    <div class="tarjeta">
      <h2>Casos por mes de registro</h2>
      <div class="sub">De los casos registrados cada mes, cuántos ya fueron atendidos y cuántos siguen abiertos</div>
      <div class="leyenda"><span class="l-abierto">Aún abierto</span><span class="l-atendido">Atendido</span></div>
      <div id="g-meses"></div>
    </div>
    <div class="tarjeta">
      <h2>Top 10 ubicaciones con más observaciones</h2>
      <div class="sub">Ubicación (edificio) — total de casos registrados</div>
      <div id="g-ubicaciones"></div>
    </div>
    <div class="tarjeta">
      <h2>Casos por empresa</h2>
      <div class="sub">Total de observaciones registradas por empresa</div>
      <div id="g-empresas"></div>
    </div>
    <div class="tarjeta">
      <h2>Gasto por empresa (S/)</h2>
      <div class="sub">Suma de los montos registrados en la columna MONTO (S/) del panel de control</div>
      <div id="g-gasto"></div>
    </div>
  </div>
</div>
<script>
window.DATOS = __DATOS__;
</script>
<script src="/dashboard.js"></script>
</body></html>"""

JS_DASHBOARD = """
(function(){
const DATOS = window.DATOS;
const MESES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"];
const abierto = r => r.estatus !== "ATENDIDO";
const monto = r => { const m = parseFloat(String(r.monto).replace(",", ".")); return isNaN(m) ? 0 : m; };
function esc(s){ const d=document.createElement("div"); d.textContent=s; return d.innerHTML; }
const fmtS = n => "S/ " + n.toLocaleString("es-PE", {minimumFractionDigits:2, maximumFractionDigits:2});

/* ---------- KPIs ---------- */
const total = DATOS.length;
const nAbiertos = DATOS.filter(abierto).length;
const nAtendidos = total - nAbiertos;
const gastoTotal = DATOS.reduce((s,r) => s + monto(r), 0);
document.getElementById("kpis").innerHTML = [
  ["Total de casos", total],
  ["Abiertos", nAbiertos],
  ["Atendidos", nAtendidos],
  ["% atendido", total ? Math.round(nAtendidos*100/total) + "%" : "-"],
  ["Gasto registrado", gastoTotal ? fmtS(gastoTotal) : "S/ 0.00"]
].map(([n,v]) => `<div class="kpi"><div class="valor">${v}</div><div class="nombre">${n}</div></div>`).join("");

/* ---------- barras apiladas por mes ---------- */
function graficoMeses(){
  const grupos = {};
  for (const r of DATOS) {
    const p = r.fecha.split(" ")[0].split("/");   // dd/mm/yyyy
    if (p.length !== 3) continue;
    const clave = p[2] + "-" + p[1];
    grupos[clave] = grupos[clave] || {ab:0, at:0};
    abierto(r) ? grupos[clave].ab++ : grupos[clave].at++;
  }
  const claves = Object.keys(grupos).sort();
  if (!claves.length) { document.getElementById("g-meses").innerHTML = "<div class='aviso-vacio'>Sin datos</div>"; return; }
  const W = 640, H = 260, mIzq = 30, mAbj = 34, mArr = 14;
  const maxV = Math.max(...claves.map(k => grupos[k].ab + grupos[k].at));
  const anchoUtil = W - mIzq - 10, altoUtil = H - mAbj - mArr;
  const paso = anchoUtil / claves.length;
  const barra = Math.min(34, paso * 0.62);
  let s = `<svg viewBox="0 0 ${W} ${H}" style="width:100%;height:auto">`;
  for (let i = 0; i <= 4; i++) {
    const v = Math.round(maxV * i / 4), y = H - mAbj - altoUtil * i / 4;
    s += `<line x1="${mIzq}" y1="${y}" x2="${W-10}" y2="${y}" stroke="#F0EEE8" stroke-width="1"/>`;
    s += `<text x="${mIzq-5}" y="${y+3.5}" font-size="10" fill="#8A877F" text-anchor="end">${v}</text>`;
  }
  claves.forEach((k, i) => {
    const g = grupos[k];
    const x = mIzq + paso * i + (paso - barra) / 2;
    const hAb = altoUtil * g.ab / maxV, hAt = altoUtil * g.at / maxV;
    const yAt = H - mAbj - hAt, yAb = yAt - (hAb ? hAb + 2 : 0);
    const [yy, mm] = k.split("-");
    const etiqueta = MESES[parseInt(mm,10)-1] + " " + yy.slice(2);
    if (g.at) s += `<rect x="${x}" y="${yAt}" width="${barra}" height="${hAt}" rx="2" fill="#0F6B5C"><title>${etiqueta}: ${g.at} atendidos</title></rect>`;
    if (g.ab) s += `<rect x="${x}" y="${yAb}" width="${barra}" height="${hAb}" rx="2" fill="#C2410C"><title>${etiqueta}: ${g.ab} aún abiertos</title></rect>`;
    s += `<text x="${x+barra/2}" y="${(g.ab||g.at ? Math.min(yAb, yAt) : H-mAbj) - 4}" font-size="10" fill="#6E6B64" text-anchor="middle">${g.ab + g.at}</text>`;
    s += `<text x="${x+barra/2}" y="${H-mAbj+14}" font-size="10" fill="#6E6B64" text-anchor="middle">${etiqueta}</text>`;
  });
  s += `<line x1="${mIzq}" y1="${H-mAbj}" x2="${W-10}" y2="${H-mAbj}" stroke="#D9D6CE" stroke-width="1"/></svg>`;
  document.getElementById("g-meses").innerHTML = s;
}

/* ---------- barras horizontales genéricas ---------- */
function barrasH(idDestino, pares, color, formato){
  if (!pares.length) { document.getElementById(idDestino).innerHTML = "<div class='aviso-vacio'>Sin datos</div>"; return; }
  const W = 640, fila = 26, mIzq = 230, mDer = 74;
  const H = pares.length * fila + 8;
  const maxV = Math.max(...pares.map(p => p[1]));
  let s = `<svg viewBox="0 0 ${W} ${H}" style="width:100%;height:auto">`;
  pares.forEach(([nombre, valor], i) => {
    const y = 4 + i * fila;
    const w = Math.max(2, (W - mIzq - mDer) * valor / maxV);
    const nom = nombre.length > 34 ? nombre.slice(0, 33) + "…" : nombre;
    s += `<text x="${mIzq-8}" y="${y+15}" font-size="11" fill="#1F1E1B" text-anchor="end">${esc(nom)}<title>${esc(nombre)}</title></text>`;
    s += `<rect x="${mIzq}" y="${y+3}" width="${w}" height="16" rx="3" fill="${color}"><title>${esc(nombre)}: ${formato(valor)}</title></rect>`;
    s += `<text x="${mIzq+w+6}" y="${y+15}" font-size="11" fill="#6E6B64">${formato(valor)}</text>`;
  });
  s += "</svg>";
  document.getElementById(idDestino).innerHTML = s;
}

function top(mapa, n){
  return Object.entries(mapa).sort((a,b) => b[1]-a[1]).slice(0, n);
}

/* top ubicaciones */
const ubic = {};
for (const r of DATOS) {
  const clave = r.ubicacion + (r.edificio ? " (" + r.edificio + ")" : "");
  ubic[clave] = (ubic[clave]||0) + 1;
}
barrasH("g-ubicaciones", top(ubic, 10), "#0F6B5C", v => v);

/* casos por empresa */
const emp = {};
for (const r of DATOS) { if (r.empresas) emp[r.empresas] = (emp[r.empresas]||0) + 1; }
barrasH("g-empresas", top(emp, 12), "#0F6B5C", v => v);

/* gasto por empresa */
const gasto = {};
for (const r of DATOS) { const m = monto(r); if (m && r.empresas) gasto[r.empresas] = (gasto[r.empresas]||0) + m; }
if (Object.keys(gasto).length) {
  barrasH("g-gasto", top(gasto, 12), "#B45309", fmtS);
} else {
  document.getElementById("g-gasto").innerHTML =
    "<div class='aviso-vacio'>Aún no hay montos registrados.<br>El administrador puede ingresar el MONTO (S/) de cada cotización en el panel de control<br>y este gráfico se llenará automáticamente.</div>";
}

graficoMeses();
})();
"""

PAGINA_HISTORIAL = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#0F6B5C">
<meta name="apple-mobile-web-app-capable" content="yes">
<link rel="icon" type="image/png" href="/icono-192.png">
<link rel="apple-touch-icon" href="/icono-192.png">
<script>if ("serviceWorker" in navigator) navigator.serviceWorker.register("/sw.js");</script>
<title>Historial de cambios - CHECK LIST SS.HH</title>
<style>
__ESTILO__
.contenedor { width:100%; margin:0 auto; }
.tarjeta { background:var(--superficie); border:1px solid var(--linea); border-radius:12px; margin-bottom:10px; padding:12px 18px; }
.cab { display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap; }
h1 { font-family:Georgia,"Times New Roman",serif; font-size:19px; font-weight:400; color:var(--tinta); }
.buscador { flex:1; min-width:200px; max-width:480px; border:1px solid var(--linea); border-radius:9px; padding:7px 14px; font-size:13px; outline:none; background:var(--superficie); color:var(--tinta); }
.buscador:focus { border-color:var(--acento); }
table { width:100%; border-collapse:collapse; font-size:12.5px; }
th { background:var(--superficie); color:var(--tinta3); font-size:10px; font-weight:600; text-transform:uppercase; letter-spacing:.08em; padding:9px 8px; text-align:left; position:sticky; top:0; white-space:nowrap; border-bottom:1.5px solid #D9D6CE; }
td { border-bottom:1px solid var(--linea2); padding:6px 8px; vertical-align:top; color:var(--tinta); line-height:1.4; }
tr:hover td { background:#FAF9F5; }
.usr { display:inline-block; border-radius:999px; padding:1px 10px; font-size:11px; font-weight:600; white-space:nowrap; border:1px solid var(--linea); }
.usr-admin { color:#0F5348; border-color:#BFD8D2; background:#F2F7F6; }
.usr-editor { color:#54524C; background:#F5F4F0; }
.usr-form { color:var(--tinta3); background:var(--superficie); }
.antes { color:var(--alerta); }
.despues { color:var(--acento); font-weight:600; }
.tabla-scroll { overflow:auto; max-height:calc(100vh - 155px); padding:0 !important; }
.vacio { padding:30px; text-align:center; color:var(--tinta2); font-size:14px; }
</style></head>
<body>
<div class="nav">
  <a class="enlace" href="/">📝 Registrar observación</a>
  <a class="enlace" href="/control">📊 Panel de control</a>
  <a class="enlace" href="/dashboard">📈 Dashboard</a>
  <span class="enlace activo">🕘 Historial de cambios</span>
</div>
<div class="contenedor">
  <div class="tarjeta cab">
    <h1>🕘 Historial de cambios — quién cambió qué y cuándo</h1>
    <input id="buscar" class="buscador" type="text" placeholder="🔎 Buscar por fila, usuario, campo o valor...">
  </div>
  <div class="tarjeta tabla-scroll">
    <table>
      <thead><tr>
        <th>Fecha y hora</th><th>Fila del caso</th><th>Usuario</th><th>Campo</th>
        <th>Valor anterior</th><th>Valor nuevo</th>
      </tr></thead>
      <tbody id="cuerpo"></tbody>
    </table>
    <div id="vacio" class="vacio" style="display:none">Aún no hay cambios registrados.</div>
  </div>
</div>
<script>
const FILAS = __FILAS__;
function esc(s){ const d=document.createElement("div"); d.textContent=s; return d.innerHTML; }
function claseUsr(u){
  if (u === "ADMINISTRADOR") return "usr usr-admin";
  if (u === "EDICION LIMITADA") return "usr usr-editor";
  return "usr usr-form";
}
function pintar(){
  const q = document.getElementById("buscar").value.toLowerCase();
  const visibles = FILAS.filter(f => !q || f.join(" ").toLowerCase().includes(q)).slice().reverse();
  document.getElementById("cuerpo").innerHTML = visibles.map(f => `<tr>
    <td>${esc(f[0])}</td><td>${esc(f[1])}</td>
    <td><span class="${claseUsr(f[2])}">${esc(f[2])}</span></td>
    <td>${esc(f[3])}</td>
    <td class="antes">${esc(f[4]) || "<i>(vacío)</i>"}</td>
    <td class="despues">${esc(f[5])}</td>
  </tr>`).join("");
  document.getElementById("vacio").style.display = visibles.length ? "none" : "block";
}
document.getElementById("buscar").addEventListener("input", pintar);
pintar();
</script>
</body></html>"""

OPCIONES_FORM = {
    "prioridad": ["ALTA", "MEDIA", "BAJA"],
    "edificio": ["EL SOL", "FAISANES"],
    "piso": ["1er", "ENTRE PISO", "2do", "3er", "4to", "5to", "6to", "7mo"],
    "empresas": ["SAMITEX INDUSTRIAL", "SAMITEX MARCA", "TEXCORP", "GLOBAL SOURCING", "PRIMATEX",
                 "COMPARTIDO ENTRE VARIAS EMPRESAS", "SINERCORP", "REMATE DE FABRICA (LUKERS)",
                 "UP GRADE (BARRIO)", "El S.A", "PANORAMA OUTSOURCING / PANORAMA CSC / SINERCORP",
                 "PANORAMA STAFF"],
    "ubicacion": ["SS.HH RECEPCION", "SS.HH PASILLO BLANCO", "DUCHA DE EMERGENCIA", "SS.HH PASILLO NEGRO",
                  "SS.HH  PLANTA", "SS.HH OFICINAS", "SS.HH AUDITORIO", "SS.HH PASILLO BLANCO POR COMEDOR",
                  "DUCHA DE  VESTUARIOS", "SS.HH  INTERIOR", "DUCTO DE  LIMPIEZA"],
    "sshh": ["SS-HH VARONES", "SS-HH DAMAS", "VESTURIO VARONES", "VESTURIO DAMAS"],
    "lavatorio": [f"LAVATORIO {i}" for i in range(1, 13)],
    "urinario": [f"URINARIO {i}" for i in range(1, 5)],
    "inodoro": [f"INODORO {i}" for i in range(1, 9)],
}

# ---------------------------------------------------------------- Servidor

class Manejador(BaseHTTPRequestHandler):
    def log_message(self, *args):
        pass

    def _responder(self, codigo, contenido, tipo="application/json"):
        cuerpo = contenido.encode("utf-8")
        self.send_response(codigo)
        self.send_header("Content-Type", f"{tipo}; charset=utf-8")
        self.send_header("Content-Length", str(len(cuerpo)))
        self.end_headers()
        self.wfile.write(cuerpo)

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            pagina = (PAGINA_FORM
                      .replace("__ESTILO__", ESTILO_BASE)
                      .replace("__OPCIONES__", json.dumps(OPCIONES_FORM, ensure_ascii=False)))
            self._responder(200, pagina, "text/html")
        elif self.path.startswith("/control"):
            try:
                registros = leer_registros()
                pagina = (PAGINA_CONTROL
                          .replace("__ESTILO__", ESTILO_BASE)
                          .replace("__DATOS__", json.dumps(registros, ensure_ascii=False))
                          .replace("__ESTADOS__", json.dumps(ESTADOS, ensure_ascii=False))
                          .replace("__OPCIONES__", json.dumps(OPCIONES_FORM, ensure_ascii=False)))
                self._responder(200, pagina, "text/html")
            except PermissionError:
                self._responder(200, "<h3 style='font-family:Arial'>⚠️ El archivo Excel está abierto en otro programa. Ciérralo y recarga la página.</h3>", "text/html")
            except Exception as e:
                self._responder(200, f"<h3 style='font-family:Arial'>⚠️ Error de conexión con la base de datos: {e}</h3>", "text/html")
        elif self.path.startswith("/dashboard.js"):
            self._responder(200, JS_DASHBOARD, "application/javascript")
        elif self.path.startswith("/dashboard"):
            try:
                registros = leer_registros()
                pagina = (PAGINA_DASHBOARD
                          .replace("__ESTILO__", ESTILO_BASE)
                          .replace("__DATOS__", json.dumps(registros, ensure_ascii=False)))
                self._responder(200, pagina, "text/html")
            except PermissionError:
                self._responder(200, "<h3 style='font-family:Arial'>⚠️ El archivo Excel está abierto en otro programa. Ciérralo y recarga la página.</h3>", "text/html")
            except Exception as e:
                self._responder(200, f"<h3 style='font-family:Arial'>⚠️ Error de conexión con la base de datos: {e}</h3>", "text/html")
        elif self.path.startswith("/historial"):
            try:
                filas = leer_historial()
                pagina = (PAGINA_HISTORIAL
                          .replace("__ESTILO__", ESTILO_BASE)
                          .replace("__FILAS__", json.dumps(filas, ensure_ascii=False)))
                self._responder(200, pagina, "text/html")
            except PermissionError:
                self._responder(200, "<h3 style='font-family:Arial'>⚠️ El archivo Excel está abierto en otro programa. Ciérralo y recarga la página.</h3>", "text/html")
            except Exception as e:
                self._responder(200, f"<h3 style='font-family:Arial'>⚠️ Error de conexión con la base de datos: {e}</h3>", "text/html")
        elif self.path.startswith("/descargar"):
            try:
                contenido = generar_resumen()
                nombre = "RESUMEN CHECK LIST SSHH " + datetime.now().strftime("%d-%m-%Y") + ".xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{nombre}"')
                self.send_header("Content-Length", str(len(contenido)))
                self.end_headers()
                self.wfile.write(contenido)
            except PermissionError:
                self._responder(200, "<h3 style='font-family:Arial'>⚠️ El archivo Excel está abierto en otro programa. Ciérralo e intenta de nuevo.</h3>", "text/html")
            except Exception as e:
                self._responder(200, f"<h3 style='font-family:Arial'>⚠️ Error de conexión con la base de datos: {e}</h3>", "text/html")
        elif self.path == "/manifest.json":
            self._responder(200, MANIFEST_JSON, "application/manifest+json")
        elif self.path == "/sw.js":
            self._responder(200, SW_JS, "application/javascript")
        elif self.path in ("/icono-192.png", "/icono-512.png"):
            icono = ICONO_192 if "192" in self.path else ICONO_512
            self.send_response(200)
            self.send_header("Content-Type", "image/png")
            self.send_header("Content-Length", str(len(icono)))
            self.send_header("Cache-Control", "public, max-age=86400")
            self.end_headers()
            self.wfile.write(icono)
        else:
            self._responder(404, "No encontrado", "text/plain")

    def _leer_json(self):
        largo = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(largo).decode("utf-8"))

    def do_POST(self):
        try:
            if self.path == "/guardar":
                fila = guardar_registro(self._leer_json())
                self._responder(200, json.dumps({"ok": True, "fila": fila}))
            elif self.path == "/verificar":
                d = self._leer_json()
                clave = d.get("clave")
                if clave == CLAVE_ADMIN:
                    self._responder(200, json.dumps({"ok": True, "rol": "admin"}))
                elif clave == CLAVE_EDITOR:
                    self._responder(200, json.dumps({"ok": True, "rol": "editor"}))
                else:
                    self._responder(200, json.dumps({"ok": False}))
            elif self.path in ("/actualizar", "/editar"):
                d = self._leer_json()
                clave = d.get("clave")
                if clave not in (CLAVE_ADMIN, CLAVE_EDITOR):
                    self._responder(200, json.dumps({"ok": False, "error": "Clave incorrecta. Activa el modo administrador."}))
                    return
                usuario = "ADMINISTRADOR" if clave == CLAVE_ADMIN else "EDICION LIMITADA"
                if self.path == "/actualizar":
                    if clave == CLAVE_EDITOR:
                        # la clave limitada solo puede tocar estas dos columnas
                        d = {k: v for k, v in d.items() if k in ("fila", "clave", "comentario_mili", "proveedor")}
                    actualizar_control(d, usuario)
                else:
                    if clave != CLAVE_ADMIN:
                        self._responder(200, json.dumps({"ok": False, "error": "Solo el administrador puede editar la observación completa."}))
                        return
                    editar_registro(d, usuario)
                self._responder(200, json.dumps({"ok": True}))
            else:
                self._responder(404, json.dumps({"ok": False, "error": "Ruta no válida"}))
        except PermissionError:
            self._responder(200, json.dumps({"ok": False, "error": "El Excel está abierto en otro programa. Ciérralo e intenta de nuevo."}))
        except Exception as e:
            self._responder(200, json.dumps({"ok": False, "error": str(e)}))


def crear_excel_si_no_existe():
    """En un despliegue nuevo (p. ej. Docker) crea la base de datos con los encabezados."""
    if os.path.exists(ARCHIVO_EXCEL):
        return
    os.makedirs(CARPETA, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    ws.append(ENCABEZADOS)
    for c in ws[1]:
        c.font = Font(name="Arial", size=10, bold=True)
    ws.freeze_panes = "A2"
    wb.save(ARCHIVO_EXCEL)
    print("Base de datos nueva creada:", ARCHIVO_EXCEL)


def main():
    if not USAR_SUPABASE:
        crear_excel_si_no_existe()
    servidor = ThreadingHTTPServer(("0.0.0.0", PUERTO), Manejador)
    url = f"http://localhost:{PUERTO}"
    print("=" * 60)
    print("  CHECK LIST SS.HH DE LA CORPORACION")
    print("=" * 60)
    print(f"  Formulario:        {url}")
    print(f"  Panel de control:  {url}/control")
    print(f"  Clave de administrador (edita todo): {CLAVE_ADMIN}")
    print(f"  Clave limitada (solo COMENTARIO FMI / COTIZACION PROVEEDOR): {CLAVE_EDITOR}")
    if USAR_SUPABASE:
        print(f"  Base de datos:     Supabase ({SUPABASE_URL})")
    else:
        print(f"  Base de datos:     {os.path.basename(ARCHIVO_EXCEL)}")
    print("  Para detener, cierra esta ventana.")
    print("=" * 60)
    if ABRIR_NAVEGADOR:
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    servidor.serve_forever()


if __name__ == "__main__":
    main()
